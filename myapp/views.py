from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login,logout
from django.contrib import messages
from django.http import HttpResponse
from datetime import date,timedelta
from .models import *
from decimal import Decimal
from datetime import datetime
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.dateparse import parse_date
from django.contrib import messages
from django.db import IntegrityError
import json
from django.db.models import Sum
from django.core.exceptions import ValidationError
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from django.utils.timezone import now
from django.utils import timezone
from django.db.models import Q
import requests
from requests.exceptions import Timeout, TooManyRedirects, RequestException, HTTPError
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.shortcuts import get_object_or_404
from datetime import date
from decimal import Decimal
from django.shortcuts import render
import urllib.parse
from datetime import date, timedelta,datetime
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db.models import Sum
import io
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
from django.db import connection
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
import re
from reportlab.platypus import Image
from django.db import transaction,IntegrityError
from django.views.decorators.http import require_POST
from dateutil.rrule import rrule, DAILY, WEEKLY, MONTHLY
from django.http import FileResponse


def generar_reporte_pdf_calendario(request):
    fecha_inicio = request.GET.get('inicio')
    fecha_fin = request.GET.get('fin')

    # Obtener ocurrencias en el rango de fechas
    ocurrencias = OcurrenciaEvento.objects.filter(
        fecha__gte=fecha_inicio,
        fecha__lte=fecha_fin
    ).select_related('evento').order_by('fecha')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer,
                          pagesize=letter,
                          rightMargin=30,
                          leftMargin=30,
                          topMargin=100,  # Aumentado para el nuevo header
                          bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    # Estilos personalizados
    titulo_style = ParagraphStyle(
        'Titulo',
        parent=styles['Heading1'],
        fontSize=16,
        leading=18,
        alignment=1,  # Centrado
        spaceAfter=6,
        fontName='Helvetica-Bold'
    )

    # Cabecera del documento mejorada
    logo_path = str(settings.BASE_DIR) + '/static/images/egatur_logo.png'

    header_data = [
        # Fila 1: Logo a la derecha
        [
            '',
            '',
            Image(logo_path, width=0.8*inch, height=0.8*inch, hAlign='RIGHT') if settings.DEBUG else ''
        ],
        # Fila 2: T√≠tulo centrado
        [
            Paragraph(
                f"<b>REPORTE DE EVENTOS</b><br/>"
                f"<font size=12>{fecha_inicio} al {fecha_fin}</font>",
                titulo_style
            )
        ],
        # Fila 3: Fecha generaci√≥n
        [
            Paragraph(f"<font size=9 color='#666666'>Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}</font>", ParagraphStyle('FechaGen', alignment=1, textColor=colors.HexColor('#666666')))
        ]
    ]

    header_table = Table(header_data, colWidths=['*', '*', 2*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (2,0), (2,0), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (2,0), (2,0), 10),
        ('ALIGN', (0,1), (-1,1), 'CENTER'),
        ('VALIGN', (0,1), (-1,1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,1), (-1,1), 12),
        ('LINEBELOW', (0,1), (-1,1), 1, colors.HexColor('#e0e0e0')),
        ('ALIGN', (0,2), (-1,2), 'CENTER'),
        ('TEXTCOLOR', (0,2), (-1,2), colors.HexColor('#666666')),
        ('BOTTOMPADDING', (0,2), (-1,2), 15),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 0.2*inch))

    # Datos de la tabla
    table_data = []
    # Encabezados
    table_data.append([
        Paragraph('<b>Evento</b>', styles['BodyText']),
        Paragraph('<b>Fecha</b>', styles['BodyText']),
        Paragraph('<b>Estado</b>', styles['BodyText']),
        Paragraph('<b>Monto (S/)</b>', styles['BodyText']),
        Paragraph('<b>Saldo Pendiente</b>', styles['BodyText'])
    ])

    # Variables para totales
    total_general = 0
    total_pagado = 0
    total_pendiente = 0

    # Llenar datos
    for o in ocurrencias:
        monto = o.evento.monto or 0
        estado = 'Pagado' if o.pagado else 'Pendiente'
        saldo = 0 if o.pagado else monto

        table_data.append([
            Paragraph(o.evento.titulo, styles['BodyText']),
            o.fecha.strftime('%d/%m/%Y'),
            estado,
            f"S/ {monto:.2f}" if o.pagado else "-",
            f"S/ {saldo:.2f}" if not o.pagado else "-"
        ])

        total_general += monto
        total_pagado += monto if o.pagado else 0
        total_pendiente += saldo

    # Fila de totales
    table_data.append([
        Paragraph('<b>TOTALES</b>', styles['BodyText']),
        '',
        '',
        Paragraph(f"<b>S/ {total_pagado:.2f}</b>", styles['BodyText']),
        Paragraph(f"<b>S/ {total_pendiente:.2f}</b>", styles['BodyText'])
    ])

    # Crear tabla
    tabla = Table(table_data, colWidths=[2.5*inch, 1.2*inch, 1.2*inch, 1.5*inch, 1.5*inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f5f5f5')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#333333')),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-2), colors.white),
        ('GRID', (0,0), (-1,-2), 1, colors.HexColor('#e0e0e0')),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f8f9fa')),
        ('BOX', (0,-1), (-1,-1), 1, colors.HexColor('#e0e0e0')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
    ]))

    elements.append(tabla)

    # Pie de p√°gina mejorado
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        canvas.drawRightString(doc.width + doc.leftMargin, 0.5*inch, f"P√°gina {doc.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)

    buffer.seek(0)
    return FileResponse(buffer,
                      as_attachment=True,
                      filename=f'reporte_eventos_{fecha_inicio}_{fecha_fin}.pdf')


def obtener_total_mes(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')

    if not mes or not anio:
        return JsonResponse({'error': 'Mes y a√±o son requeridos'}, status=400)

    try:
        mes, anio = int(mes), int(anio)

        # üîπ Filtrar ocurrencias del mes y que NO est√©n pagadas
        total = OcurrenciaEvento.objects.filter(
            fecha__year=anio,
            fecha__month=mes,
            pagado=False  # Solo eventos NO pagados
        ).aggregate(total=Sum('evento__monto'))['total'] or 0

        return JsonResponse({'total': round(total, 2)})

    except ValueError:
        return JsonResponse({'error': 'Mes y a√±o deben ser num√©ricos'}, status=400)


@login_required
@require_POST
def actualizar_evento(request):
    try:
        data = json.loads(request.body)
        event_id = data.get("id")
        nueva_fecha = data.get("fecha")

        evento = OcurrenciaEvento.objects.get(id=event_id)
        evento.fecha = nueva_fecha
        evento.save()

        return JsonResponse({"success": "Fecha actualizada correctamente."})
    except OcurrenciaEvento.DoesNotExist:
        return JsonResponse({"error": "Evento no encontrado."}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"Ocurri√≥ un error: {str(e)}"}, status=500)


@login_required
def eliminar_ocurrencia_evento(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            occ_id = data.get('id')
            print("ID de ocurrencia a eliminar:", occ_id)
            ocurrencia = OcurrenciaEvento.objects.get(id=occ_id)
            ocurrencia.delete()
            return JsonResponse({'success': True, 'message': 'La ocurrencia se ha eliminado correctamente.'})
        except OcurrenciaEvento.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Ocurrencia no encontrada.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'error': 'M√©todo no permitido'}, status=405)
@login_required
def pagar_evento(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            occ_id = data.get('id')  # Recibe el ID del evento
            evento = OcurrenciaEvento.objects.get(id=occ_id)

            evento.pagado = True  # Marcar como pagado
            evento.save()

            return JsonResponse({'success': True, 'message': 'El evento ha sido marcado como pagado.'})
        except Evento.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Evento no encontrado.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'error': 'M√©todo no permitido.'}, status=405)


@login_required
def gasto_calendario(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            fecha = data.get('fecha')
            importe = data.get('importe')
            metodo_pago = data.get('metodo_pago')
            moneda = data.get('moneda')
            local_id = data.get('local')
            tipo_comprobante = data.get('tipo_comprobante')
            nombre_proveedor = data.get('nombre_proveedor')
            observacion = data.get('observacion')
            codigo_operacion = data.get('codigo_operacion') if metodo_pago != 'efectivo' else None
            fecha_operacion = data.get('fecha_operacion') if metodo_pago != 'efectivo' else None
            concepto_nivel_1 = data.get('concepto_nivel_1')
            concepto_nivel_2 = data.get('concepto_nivel_2')
            concepto_nivel_3 = data.get('concepto_nivel_3')
            numero_comprobante = data.get('num_comprobante')
            fecha_emision_comprobante = data.get('fecha_emision_comprobante') if tipo_comprobante in ['RHE', 'Factura', 'Boleta', 'Nota', 'Proforma'] else None
            campo_area = data.get('campo_area')
            campo_mes = data.get('campo_mes') if tipo_comprobante == 'Boleta de pago' else None
            id_requerimiento = data.get('id_requerimiento') if tipo_comprobante == 'Requerimiento' else None
            num_requerimiento = data.get('num_requerimiento') if tipo_comprobante == 'Requerimiento' else None
            banco_id = data.get('banco_operacion') if metodo_pago != 'efectivo' else None
            occ_id = data.get('event_id')

            # Validaci√≥n de campos obligatorios
            if not all([fecha, importe, metodo_pago, moneda]):
                return JsonResponse({'error': 'Todos los campos obligatorios deben completarse.'}, status=400)

            # Validaci√≥n de tipo de comprobante
            if tipo_comprobante in ['RHE', 'Factura', 'Boleta'] and (not numero_comprobante or not fecha_emision_comprobante):
                return JsonResponse({'error': 'N√∫mero y Fecha de Emisi√≥n del Comprobante son obligatorios.'}, status=400)
            if tipo_comprobante == 'Boleta de pago' and not campo_mes:
                return JsonResponse({'error': 'El campo "Mes" es obligatorio para "Boleta de pago".'}, status=400)
            if tipo_comprobante == 'Requerimiento' and not id_requerimiento:
                return JsonResponse({'error': 'El ID Requerimiento es obligatorio para "Requerimiento".'}, status=400)

            # Validar formato de fecha
            try:
                fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha inv√°lido (YYYY-MM-DD requerido).'}, status=400)

            # Buscar proveedor
            proveedor = Proveedor.objects.filter(id=nombre_proveedor).first()
            if not proveedor:
                return JsonResponse({'error': 'El proveedor especificado no existe.'}, status=404)

            # Obtener el local
            local = Local.objects.filter(id=local_id).first() if local_id else None
            if local_id and not local:
                return JsonResponse({'error': 'El local especificado no existe.'}, status=404)

            # Validar conceptos
            resultado_conceptos = comprobar_conceptos(tipo_comprobante, concepto_nivel_1, concepto_nivel_2, concepto_nivel_3)
            if 'error' in resultado_conceptos:
                return JsonResponse({'error': 'Revise los niveles de los conceptos, falta llenar algunos campos.'}, status=400)

            concepto_1, concepto_2, concepto_3 = resultado_conceptos['concepto_nivel_1'], resultado_conceptos['concepto_nivel_2'], resultado_conceptos['concepto_nivel_3']

            # Obtener banco si se proporciona
            banco = Banco.objects.filter(id=banco_id).first() if banco_id else None
            if banco_id and not banco:
                return JsonResponse({'error': 'El banco especificado no existe.'}, status=404)

            # Obtener evento
            evento = OcurrenciaEvento.objects.filter(id=occ_id).first()
            if not evento:
                return JsonResponse({'error': 'El evento especificado no existe.'}, status=404)

            # Transacci√≥n at√≥mica: si algo falla, nada se guarda
            with transaction.atomic():
                # Crear gasto
                gasto = Gasto.objects.create(
                    fecha_gasto=fecha,
                    concepto_nivel_1=concepto_1,
                    concepto_nivel_2=concepto_2,
                    concepto_nivel_3=concepto_3,
                    importe=importe,
                    nombre_proveedor=proveedor,
                    local=local,
                    tipo_comprobante=tipo_comprobante,
                    tipo_pago=metodo_pago,
                    codigo_operacion=codigo_operacion,
                    observacion=observacion,
                    fecha_operacion=fecha_operacion,
                    moneda=moneda,
                    numero_comprobante=numero_comprobante,
                    fecha_emision_comprobante=fecha_emision_comprobante,
                    campo_area=campo_area,
                    campo_mes=campo_mes,
                    id_requerimiento=id_requerimiento,
                    num_requerimiento=num_requerimiento,
                    banco=banco,
                    usuario_creador=request.user
                )

                # Marcar el evento como pagado
                evento.pagado = True
                evento.save()

            return JsonResponse({'success': 'Gasto registrado correctamente.'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Error al procesar la solicitud. Datos JSON inv√°lidos.'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'Error inesperado: {str(e)}'}, status=500)


@login_required
def eliminar_evento(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            event_id = data.get('id')

            if not event_id:
                return JsonResponse({'success': False, 'error': 'ID del evento no proporcionado.'}, status=400)

            evento = Evento.objects.get(id=event_id, creado_por=request.user)

            # Eliminar todas las ocurrencias relacionadas
            OcurrenciaEvento.objects.filter(evento=evento).delete()

            # Eliminar el evento maestro
            evento.delete()

            return JsonResponse({'success': True, 'message': 'El evento y sus ocurrencias han sido eliminados correctamente.'})

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Error en el formato de la solicitud.'}, status=400)
        except Evento.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Evento no encontrado o no tienes permisos para eliminarlo.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'error': 'M√©todo no permitido.'}, status=405)
from django.db.models import Sum, Count, Case, When, DecimalField

@login_required
def resumen_eventos(request):

    resumen = (
        OcurrenciaEvento.objects
        .annotate(month=TruncMonth('fecha'))
        .values('month')
        .annotate(
            total=Sum('evento__monto', output_field=DecimalField()),
            pendientes=Count(Case(When(pagado=False, then=1))),
            pagos=Count(Case(When(pagado=True, then=1)))
        )
        .order_by('month')
    )
    data = []
    for item in resumen:
        month_date = item['month']
        total = item['total'] or 0
        pendientes = item['pendientes']
        pagos = item['pagos']
        # Aqu√≠ se usa strftime y luego se pasa a may√∫sculas; si tu servidor tiene locale configurado en espa√±ol se mostrar√° correctamente.
        month_str = month_date.strftime("%B %Y").upper()
        data.append({
            'month': month_str,
            'total': str(total),  # Puedes formatearlo como moneda si lo deseas
            'pendientes': pendientes,
            'pagos': pagos,
        })
    return JsonResponse(data, safe=False)
@login_required
def obtener_eventos_pagados(request):
    # Filtramos solo ocurrencias pagadas
    ocurrencias = OcurrenciaEvento.objects.filter(pagado=True)
    events_list = []

    for occ in ocurrencias:
        evento = occ.evento
        start_date = occ.fecha.isoformat()  # Usamos la fecha espec√≠fica de la ocurrencia

        event_dict = {
            'id': f"pagado_{occ.id}",  # Prefijo para identificar
            'title': f"{evento.titulo}",  # Checkmark para indicar pagado
            'start': start_date,
            'allDay': True,
            'extendedProps': {
                'monto': str(evento.monto) if evento.monto is not None else "",
                'notas': evento.notas,
                'recurrencia': evento.recurrencia,
                'pagado': True,  # Siempre true en este endpoint
                'evento_id': evento.id,
                'prestamo': evento.prestamo,
                'ocurrencia_id': occ.id  # ID espec√≠fico de la ocurrencia
            }
        }

        events_list.append(event_dict)

    return JsonResponse(events_list, safe=False)

@login_required
def obtener_eventos(request):
    # Se obtienen todas las ocurrencias
    ocurrencias = OcurrenciaEvento.objects.filter(pagado=False)
    events_list = []

    for occ in ocurrencias:
        evento = occ.evento
        start_date = evento.fecha_inicio.isoformat()
        end_date = None  # Inicializar end_date

        if evento.recurrencia.lower() == 'none':
            if evento.fecha_fin and evento.fecha_fin != evento.fecha_inicio:
                end_date = (evento.fecha_fin + timedelta(days=1)).isoformat()  # üîπ Sumar 1 d√≠a a end_date
            event_dict = {
                'id': occ.id,
                'title': evento.titulo,
                'start': start_date,
                'allDay': True,
                'extendedProps': {
                    'monto': str(evento.monto) if evento.monto is not None else "",
                    'notas': evento.notas,
                    'recurrencia': evento.recurrencia,
                    'pagado': occ.pagado,
                    'evento_id': evento.id,
                    'prestamo': evento.prestamo  # üîπ Agregar info de pr√©stamo

                }
            }

            if end_date:  # ‚úÖ Asegurar que 'end' se agregue correctamente
                event_dict['end'] = end_date

        else:
            # Para eventos recurrentes, cada ocurrencia se muestra individualmente
            event_dict = {
                'id': occ.id,
                'title': evento.titulo,
                'start': occ.fecha.isoformat(),
                'allDay': True,
                'extendedProps': {
                    'monto': str(evento.monto) if evento.monto is not None else "",
                    'notas': evento.notas,
                    'recurrencia': evento.recurrencia,
                    'pagado': occ.pagado,
                    'evento_id': evento.id,
                    'prestamo': evento.prestamo  # üîπ Agregar info de pr√©stamo

                }
            }

        events_list.append(event_dict)

    return JsonResponse(events_list, safe=False)


@login_required
def crear_evento(request):
    if request.method == 'POST':
        try:
            print("Recibiendo solicitud para crear evento...")
            data = json.loads(request.body)
            print("Datos recibidos:", data)

            titulo = data.get('titulo')
            fecha_inicio = data.get('fecha_inicio')
            fecha_fin = data.get('fecha_fin')  # Puede ser None o vac√≠o
            recurrencia = data.get('recurrencia', 'none').lower()
            monto = data.get('monto')
            notas = data.get('notas')
            prestamo = data.get('prestamo', False)  # Valor booleano
            repeat_until = data.get('repeatUntil')

            # Validaci√≥n b√°sica
            if not titulo or not fecha_inicio:
                print("Error: Faltan campos obligatorios.")
                return JsonResponse({'error': 'Faltan campos obligatorios'}, status=400)

            # Para eventos recurrentes, si se env√≠a repeatUntil, usarlo como fecha_fin
            if recurrencia != 'none' and repeat_until:
                print("Evento recurrente: usando 'repeatUntil' como fecha_fin:", repeat_until)
                fecha_fin = repeat_until

            print("Creando evento maestro...")
            evento = Evento.objects.create(
                titulo=titulo,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                recurrencia=recurrencia,
                monto=monto if monto != "" else None,
                notas=notas,
                creado_por=request.user,
                prestamo=prestamo
            )
            print("Evento maestro creado:", evento)

            # Creaci√≥n de ocurrencias:
            if recurrencia == 'none':
                # Evento sin recurrencia: se crea una √∫nica ocurrencia en fecha_inicio.
                dtstart = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                if fecha_fin and fecha_fin != fecha_inicio:
                    dtend = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                    print(f"Creando √∫nica ocurrencia para evento de varios d√≠as: {dtstart} - {dtend}")
                    OcurrenciaEvento.objects.create(evento=evento, fecha=dtstart)
                else:
                    print(f"Creando √∫nica ocurrencia para evento de un d√≠a: {dtstart}")
                    OcurrenciaEvento.objects.create(evento=evento, fecha=dtstart)
            else:
                # Evento recurrente: se generan m√∫ltiples ocurrencias.
                if fecha_fin:
                    dtstart = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                    dtend = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                    print("Evento recurrente. dtstart:", dtstart, "dtend:", dtend)
                    freq = {'daily': DAILY, 'weekly': WEEKLY, 'monthly': MONTHLY}.get(recurrencia, None)
                    if freq:
                        print("Generando ocurrencias con frecuencia:", recurrencia)
                        for occ in rrule(freq, dtstart=dtstart, until=dtend):
                            occ_date = occ.date()
                            OcurrenciaEvento.objects.create(evento=evento, fecha=occ_date)
                            print("Creada ocurrencia:", occ_date)
                else:
                    dtstart = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                    print("Evento recurrente sin l√≠mite. dtstart:", dtstart)
                    OcurrenciaEvento.objects.create(evento=evento, fecha=dtstart)
                    print("Creada √∫nica ocurrencia para evento recurrente sin l√≠mite:", dtstart)

            print("Evento y ocurrencias creados correctamente.")
            return JsonResponse({
                'success': True,
                'id': evento.id,
                'titulo': evento.titulo
            })
        except Exception as e:
            print("Error en crear_evento:", e)
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'M√©todo no permitido'}, status=405)

def ver_calendar(request):
    nivel_1_conceptos = Concepto.objects.filter(nivel=1)
    nivel_2_conceptos = Concepto.objects.filter(nivel=2)
    nivel_3_conceptos = Concepto.objects.filter(nivel=3)
    fondos = Fondo.objects.all()
    locales = Local.objects.all()
    bancos = Banco.objects.all()
    proveedores = Proveedor.objects.all()  # Obtener todos los proveedores
    nivel_1_conceptos_json = json.dumps(list(nivel_1_conceptos.values_list('id', 'concepto_nombre')), default=str)

    return render(request, 'calendario.html', {
        'nivel_1_conceptos_json': nivel_1_conceptos_json,
        'fondos': fondos,
        'locales': locales,
        'proveedores': proveedores,  # Pasar los proveedores al contexto
        'nivel_1_conceptos': nivel_1_conceptos,
        'nivel_2_conceptos': nivel_2_conceptos,
        'nivel_3_conceptos': nivel_3_conceptos,
        'bancos':bancos
    })


# ------------------------------------------------------



def generar_pdf_rendiciones(request):
    fecha_desde = request.GET.get("desde", "")
    fecha_hasta = request.GET.get("hasta", "")

    # Convertir a formato datetime para filtrar correctamente
    fecha_inicio = datetime.strptime(fecha_desde, "%Y-%m-%d")
    fecha_fin = datetime.strptime(fecha_hasta, "%Y-%m-%d")

    # Filtrar los gastos por fecha_rendido dentro del rango dado (INCLUYENDO el 25 y 27)
    gastos = Gasto.objects.filter(fecha_rendido__gte=fecha_inicio, fecha_rendido__lte=fecha_fin)

    # Nombre del archivo PDF
    nombre_archivo = f"Rendiciones_{fecha_desde}_a_{fecha_hasta}.pdf"

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'

    # Crear el documento PDF
    doc = SimpleDocTemplate(response, pagesize=A4, leftMargin=inch, rightMargin=inch, topMargin=inch, bottomMargin=inch)
    elements = []
    styles = getSampleStyleSheet()

    # Estilos personalizados
    estilo_titulo = ParagraphStyle("Titulo", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#255a9e"), alignment=1)
    estilo_texto = ParagraphStyle("Normal", parent=styles["BodyText"], fontSize=10)
    estilo_header = ParagraphStyle("Header", parent=styles["BodyText"], fontSize=12, textColor=colors.white)

    # üè∑Ô∏è **Portada del PDF**
    elements.append(Spacer(1, 80))  # Aumentar el espacio antes del t√≠tulo
    elements.append(Paragraph("<b>RENDICIONES</b>", estilo_titulo))
    elements.append(Spacer(1, 30))  # Aumentar el espacio entre el t√≠tulo y la fecha
    elements.append(Paragraph(f"Desde: <b>{fecha_desde}</b> Hasta: <b>{fecha_hasta}</b>", estilo_titulo))
    elements.append(Spacer(1, 100))  # Aumentar espacio antes de la imagen

    # üìå **Imagen centrada m√°s abajo**
    logo_path = f"{settings.STATICFILES_DIRS[0]}/images/egatur_logo.png"  # Ajusta la ruta si es necesario
    img = Image(logo_path, width=380, height=380)  # Ajusta el tama√±o si es necesario
    img.hAlign = 'CENTER'  # Centra la imagen
    elements.append(img)

    elements.append(PageBreak())  # Salto de p√°gina



    # üìú **Procesar cada gasto**
    for gasto in gastos:
        rendiciones = gasto.rendiciones_gasto.all()
        total_importe = sum(r.importe for r in rendiciones if r.importe)

        # üîπ **Encabezado del gasto**
        elements.append(Paragraph(f"<b>Rendici√≥n de Gasto #{gasto.id}</b>", estilo_titulo))
        elements.append(Spacer(1, 12))

        # üîπ **Datos del gasto**
        datos_gasto = [
            [Paragraph("<b>ID Requerimiento:</b>", estilo_texto), Paragraph(gasto.id_requerimiento or "-", estilo_texto),
             Paragraph("<b>N¬∞ Requerimiento:</b>", estilo_texto), Paragraph(gasto.num_requerimiento or "-", estilo_texto)],
            [Paragraph("<b>√Årea:</b>", estilo_texto), Paragraph(gasto.campo_area or "-", estilo_texto),
             Paragraph("<b>Fecha de Gasto:</b>", estilo_texto), Paragraph(gasto.fecha_gasto.strftime('%d/%m/%Y') if gasto.fecha_gasto else "-", estilo_texto)],
            [Paragraph("<b>Proveedor:</b>", estilo_texto), Paragraph(gasto.nombre_proveedor.razon_social if gasto.nombre_proveedor else "-", estilo_texto),
             Paragraph("<b>Local:</b>", estilo_texto), Paragraph(gasto.local.nombre_local if gasto.local else "-", estilo_texto)],
            [Paragraph("<b>Importe Total:</b>", estilo_texto), Paragraph(f"S/. {gasto.importe:,.2f} {gasto.moneda}", estilo_texto),
             Paragraph("<b>Observaci√≥n:</b>", estilo_texto), Paragraph(gasto.observacion or "-", estilo_texto)]
        ]

        tabla_gasto = Table(datos_gasto, colWidths=[100, 180, 100, 180])
        tabla_gasto.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID", (0, 0), (-1, -1), 0, colors.white),  # Bordes invisibles
            ("WORDWRAP", (0, 0), (-1, -1), "CJK"),  # Evita desbordes de texto
        ]))

        elements.append(tabla_gasto)
        elements.append(Spacer(1, 12))

        # üîπ **Detalle de Rendiciones**
        if rendiciones:
            elements.append(Paragraph("<b>Detalle de Rendiciones</b>", styles["Heading2"]))
            elements.append(Spacer(1, 10))

            datos_rendiciones = [
                [Paragraph("<b>Fecha</b>", estilo_header),
                 Paragraph("<b>Descripci√≥n</b>", estilo_header),
                 Paragraph("<b>N¬∞ Requerimiento</b>", estilo_header),
                 Paragraph("<b>Tipo Comprobante</b>", estilo_header),
                 Paragraph("<b>Proveedor</b>", estilo_header),
                 Paragraph("<b>Importe</b>", estilo_header)]
            ]

            for rendicion in rendiciones:
                datos_rendiciones.append([
                    Paragraph(rendicion.fecha_operacion.strftime('%d/%m/%Y') if rendicion.fecha_operacion else "-", estilo_texto),
                    Paragraph(rendicion.descripcion or "-", estilo_texto),
                    Paragraph(rendicion.numero_requerimiento or "-", estilo_texto),
                    Paragraph(rendicion.tipo_comprobante or "-", estilo_texto),
                    Paragraph(rendicion.proveedor.razon_social if rendicion.proveedor else "-", estilo_texto),
                    Paragraph(f"S/. {rendicion.importe:,.2f}", estilo_texto)
                ])

            # Agregar fila de total
            datos_rendiciones.append(["", "", "", "", Paragraph("<b>Total:</b>", estilo_texto), Paragraph(f"S/. {total_importe:,.2f}", estilo_texto)])

            tabla_rendiciones = Table(datos_rendiciones, colWidths=[70, 150, 80, 100, 100, 80])
            tabla_rendiciones.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#255a9e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))

            elements.append(tabla_rendiciones)
            elements.append(Spacer(1, 12))

        # üîπ **Diferencia entre gasto y rendiciones**
        diferencia = gasto.importe - total_importe
        if diferencia < 0:
            mensaje = Paragraph(f"<b>Se hizo un gasto extra de: S/. {abs(diferencia):,.2f}</b>", estilo_texto)
        elif diferencia > 0:
            mensaje = Paragraph(f"<b>Se hizo un ingreso de: S/. {diferencia:,.2f}</b>", estilo_texto)
        else:
            mensaje = Paragraph("<b></b>", estilo_texto)

        elements.append(mensaje)
        elements.append(PageBreak())

    # üìÑ **Construcci√≥n final del PDF**
    doc.build(elements)

    return response

def limpiar_nombre_archivo(texto):
    return re.sub(r'[\\/*?:"<>|]', '', texto)

def crear_pdf_rendicion(request, gasto_id):
    gasto = get_object_or_404(Gasto, id=gasto_id)
    rendiciones = gasto.rendiciones_gasto.all()

    nombre_archivo = f"Requerimiento_{gasto.id_requerimiento}_N_{gasto.num_requerimiento}.pdf"
    nombre_archivo = limpiar_nombre_archivo(nombre_archivo)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'

    doc = SimpleDocTemplate(response, pagesize=A4, leftMargin=inch, rightMargin=inch, topMargin=inch, bottomMargin=inch)
    elements = []
    styles = getSampleStyleSheet()

    # Estilos personalizados
    estilo_titulo = ParagraphStyle("Titulo", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#255a9e"))
    estilo_texto = ParagraphStyle("Normal", parent=styles["BodyText"], fontSize=9)  # Tama√±o reducido para evitar desbordes
    estilo_header = ParagraphStyle("Header", parent=styles["BodyText"], fontSize=12, textColor=colors.white)
    estilo_alerta_rojo = ParagraphStyle("AlertaRojo", parent=styles["BodyText"], fontSize=11, textColor=colors.HexColor("#D32F2F"))  # Rojo pastel
    estilo_alerta_verde = ParagraphStyle("AlertaVerde", parent=styles["BodyText"], fontSize=11, textColor=colors.HexColor("#388E3C"))  # Verde suave

    # T√≠tulo del documento
    elements.append(Paragraph(f"<b>Rendici√≥n de Gasto #{gasto.id}</b>", estilo_titulo))
    elements.append(Spacer(1, 12))

    # Datos del gasto en dos columnas
    datos_gasto = [
        [Paragraph("<b>ID Requerimiento:</b>", estilo_texto), Paragraph(gasto.id_requerimiento or "-", estilo_texto),
         Paragraph("<b>N¬∞ Requerimiento:</b>", estilo_texto), Paragraph(gasto.num_requerimiento or "-", estilo_texto)],
        [Paragraph("<b>√Årea:</b>", estilo_texto), Paragraph(gasto.campo_area or "-", estilo_texto),
         Paragraph("<b>Fecha de Gasto:</b>", estilo_texto), Paragraph(gasto.fecha_gasto.strftime('%d/%m/%Y') if gasto.fecha_gasto else "-", estilo_texto)],
        [Paragraph("<b>Proveedor:</b>", estilo_texto), Paragraph(gasto.nombre_proveedor.razon_social if gasto.nombre_proveedor else "-", estilo_texto),
         Paragraph("<b>Local:</b>", estilo_texto), Paragraph(gasto.local.nombre_local if gasto.local else "-", estilo_texto)],
        [Paragraph("<b>Importe Total:</b>", estilo_texto), Paragraph(f"S/. {gasto.importe:,.2f} {gasto.moneda}", estilo_texto),
         Paragraph("<b>Observaci√≥n:</b>", estilo_texto), Paragraph(gasto.observacion or "-", estilo_texto)]
    ]

    tabla_gasto = Table(datos_gasto, colWidths=[100, 180, 100, 180])
    tabla_gasto.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0, colors.white),  # Bordes invisibles
        ("WORDWRAP", (0, 0), (-1, -1), "CJK"),  # Evita desbordes de texto
    ]))

    elements.append(tabla_gasto)
    elements.append(Spacer(1, 12))

    # Detalle de rendiciones
    if rendiciones:
        elements.append(Paragraph("<b>Detalle de Rendiciones</b>", styles["Heading2"]))
        elements.append(Spacer(1, 10))

        datos_rendiciones = [
            [Paragraph("<b>Fecha</b>", estilo_header),
             Paragraph("<b>Descripci√≥n</b>", estilo_header),
             Paragraph("<b>N¬∞ Requerimiento</b>", estilo_header),
             Paragraph("<b>Tipo Comprobante</b>", estilo_header),
             Paragraph("<b>Proveedor</b>", estilo_header),
             Paragraph("<b>Importe</b>", estilo_header)]
        ]

        total_importe = 0
        for rendicion in rendiciones:
            importe = rendicion.importe if rendicion.importe else 0
            total_importe += importe
            datos_rendiciones.append([
                Paragraph(rendicion.fecha_operacion.strftime('%d/%m/%Y') if rendicion.fecha_operacion else "-", estilo_texto),
                Paragraph(rendicion.descripcion or "-", estilo_texto),
                Paragraph(rendicion.numero_requerimiento or "-", estilo_texto),
                Paragraph(rendicion.tipo_comprobante or "-", estilo_texto),
                Paragraph(rendicion.proveedor.razon_social if rendicion.proveedor else "-", estilo_texto),
                Paragraph(f"S/. {importe:,.2f}", estilo_texto)
            ])

        # Agregar fila de total
        datos_rendiciones.append(["", "", "", "", Paragraph("<b>Total:</b>", estilo_texto), Paragraph(f"S/. {total_importe:,.2f}", estilo_texto)])

        # Crear la tabla
        tabla_rendiciones = Table(datos_rendiciones, colWidths=[70, 150, 80, 100, 100, 80])
        tabla_rendiciones.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#255a9e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("BACKGROUND", (-1, -1), (-1, -1), colors.lightgrey),  # Resaltar Total
            ("WORDWRAP", (0, 0), (-1, -1), "CJK"),  # Evita desbordes de texto
        ]))

        elements.append(tabla_rendiciones)
    # Calcular diferencia
    diferencia = gasto.importe - total_importe

    # Mostrar mensaje en rojo o verde dependiendo del valor de la diferencia
    if diferencia < 0:
        mensaje = Paragraph(f"<b>Se hizo un gasto extra de: S/. {abs(diferencia):,.2f}</b>", estilo_alerta_rojo)
        elements.append(mensaje)
    elif diferencia > 0:
        mensaje = Paragraph(f"<b>Se hizo un ingreso de: S/. {diferencia:,.2f}</b>", estilo_alerta_verde)
        elements.append(mensaje)
    # Construir el PDF
    doc.build(elements)

    return response

@login_required
def ver_rendidos(request):
    # Filtra solo los gastos que han sido rendidos (rendido=True)
    gastos_rendiciones = Gasto.objects.filter(rendido=True)
    return render(request, 'ver_rendidos.html', {'rendiciones': gastos_rendiciones})

@login_required
def ver_rendiciones_asociadas(request, gasto_id):
    # Obtener el gasto o devolver 404 si no existe
    gasto = get_object_or_404(Gasto, id=gasto_id)

    # Obtener todas las rendiciones asociadas a este gasto
    rendiciones = gasto.rendiciones_gasto.all()

    # Convertir los datos a JSON
    data = []
    for rendicion in rendiciones:
        data.append({
            'id': rendicion.id,
            'fecha_operacion': rendicion.fecha_operacion.strftime('%d/%m/%Y') if rendicion.fecha_operacion else None,
            'descripcion': rendicion.descripcion or '',
            'numero_requerimiento': rendicion.numero_requerimiento or '',
            'importe': float(rendicion.importe) if rendicion.importe else 0,
            'tipo_comprobante': rendicion.tipo_comprobante or '',
            'proveedor': rendicion.proveedor.razon_social if rendicion.proveedor else '',
        })

    # Retornar los datos en formato JSON
    return JsonResponse({'rendiciones': data}, safe=False)


def reiniciar_secuencia(request):
    tablas = [
        Gasto._meta.db_table,
        Ingreso._meta.db_table,
        Rendicion._meta.db_table,
    ]

    with connection.cursor() as cursor:
        for tabla in tablas:
            # Borrar todos los registros de la tabla
            cursor.execute(f"DELETE FROM {tabla};")
            # Reiniciar la secuencia en SQLite
            cursor.execute(f"DELETE FROM sqlite_sequence WHERE name='{tabla}';")

    return HttpResponse("Tablas limpiadas y secuencias reiniciadas.")

@csrf_exempt
def registrar_usuario(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            username = data["username"]
            password = data["password"]
            saldo_efectivo = data["saldo_efectivo"]
            saldo_yape = data["saldo_yape"]

            if User.objects.filter(username=username).exists():
                return JsonResponse({"success": False, "message": "El usuario ya existe."})

            usuario = User.objects.create_user(username=username, password=password, is_active=True, is_staff=False)
            SaldoInicial.objects.create(usuario=usuario, monto_saldo_inicial=saldo_efectivo, monto_saldo_inicial_yape=saldo_yape)

            return JsonResponse({"success": True, "message": "Usuario registrado correctamente."})
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Error: {str(e)}"})

    return JsonResponse({"success": False, "message": "M√©todo no permitido."})
def cerrar_caja(request):
    usuarios_no_staff = User.objects.filter(is_staff=False)  # Obtener usuarios normales
    return render(request, 'cerrarcaja.html', {'usuarios': usuarios_no_staff})
@login_required
def cerrar_caja_usuario(request, user_id):
    if request.method == "POST":
        usuario = get_object_or_404(User, id=user_id)
        saldo = get_object_or_404(SaldoInicial, usuario=usuario)

        if not saldo.caja_cerrada:
            saldo.caja_cerrada = True
            saldo.fecha_cierre = now().date()
            saldo.save()
            return JsonResponse({"success": True, "message": f"La caja de {usuario.username} ha sido cerrada."})
        else:
            return JsonResponse({"success": False, "message": "La caja ya est√° cerrada."})

    return JsonResponse({"success": False, "message": "M√©todo no permitido."}, status=405)

@login_required
def reactivar_caja_usuario(request, user_id):
    if request.method == "POST":
        usuario = get_object_or_404(User, id=user_id)
        saldo = get_object_or_404(SaldoInicial, usuario=usuario)

        if saldo.caja_cerrada:
            saldo.caja_cerrada = False
            saldo.fecha_cierre = None
            saldo.save()
            return JsonResponse({"success": True, "message": f"La caja de {usuario.username} ha sido reactivada."})
        else:
            return JsonResponse({"success": False, "message": "La caja ya est√° activa."})

    return JsonResponse({"success": False, "message": "M√©todo no permitido."}, status=405)


def editar_personal(request, id):
    personal = get_object_or_404(Personal, id=id)
    bancos=Banco.objects.all()
    locales = Local.objects.all()

    # Convertir remuneraci√≥n a string con dos decimales si no es None
    if personal.remuneracion is not None:
        personal.remuneracion = f"{float(personal.remuneracion):.2f}"
    return render(request, 'editar_personal.html', {'personal': personal ,'bancos':bancos , 'locales':locales})

def crear_contrase√±a(request, personal_id):
    personal = get_object_or_404(Personal, id=personal_id)

    # Datos a enviar al otro sistema
    nombre = personal.apellidos_nombres
    telefono = personal.celular

    # URL del otro sistema que recibe los datos
    url_crear_cliente = "http://cafeteria.egatur.edu.pe/crear-cliente/"

    # Par√°metros para la petici√≥n GET
    params = {
        "nombre": nombre,
        "telefono": telefono
    }

    try:
        response = requests.get(url_crear_cliente, params=params, timeout=5)

        # Verifica si la respuesta tiene contenido
        if response.status_code == 201:
            messages.success(request, "Cliente creado exitosamente en el otro sistema.")
        elif response.status_code == 204:
            # Si la respuesta es 204 (sin contenido), se puede mostrar un mensaje
            messages.warning(request, "La respuesta del sistema est√° vac√≠a, pero el cliente no present√≥ errores.")
        else:
            # Verifica si la respuesta tiene un cuerpo JSON
            try:
                response_data = response.json()
                error_message = response_data.get("mensaje", "No se pudo crear el cliente.")
                messages.warning(request, f"Error al crear el cliente: {error_message} (C√≥digo de estado: {response.status_code})")
            except ValueError:
                # Si la respuesta no es un JSON v√°lido, muestra el contenido de la respuesta como texto
                messages.error(request, f"Error al procesar la respuesta del sistema: {response.text} (C√≥digo de estado: {response.status_code})")

    # Manejo de diferentes excepciones
    except Timeout:
        messages.error(request, "Error: Tiempo de espera agotado al conectar con el otro sistema.")

    except TooManyRedirects:
        messages.error(request, "Error: Demasiados redireccionamientos al conectar con el otro sistema.")

    except HTTPError as http_err:
        messages.error(request, f"Error HTTP al conectar con el otro sistema: {http_err}")

    except RequestException as e:
        # Cualquier otra excepci√≥n gen√©rica de requests
        messages.error(request, f"Error inesperado al conectar con el otro sistema: {str(e)}")

    return redirect("ver_personal")
from reportlab.lib import colors
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import os
from django.conf import settings

def draw_centered_text(p, text, y_position, font="Helvetica-Bold", size=16, page_width=letter[0]):
    p.setFont(font, size)
    text_width = p.stringWidth(text, font, size)
    x_position = (page_width - text_width) / 2  # Centrar el texto
    p.drawString(x_position, y_position, text)

def draw_logo(p, image_name, x=490, y=690, width=75, height=75):
    image_path = os.path.join(settings.BASE_DIR, "static/images", image_name)

    if os.path.exists(image_path):  # Verifica si la imagen existe
        p.drawImage(image_path, x, y, width, height, mask='auto')
    else:
        print(f"‚ö†Ô∏è Error: No se encontr√≥ la imagen en {image_path}")

def generar_pdf(request, persona_id):
    try:
        persona = Personal.objects.get(id=persona_id)

        # Crear la respuesta como un archivo PDF
        response = HttpResponse(content_type='application/pdf')
        filename = f'Ficha_{persona.apellidos_nombres}.pdf'.replace(" ", "_")  # Reemplaza espacios por guiones bajos
        filename = urllib.parse.quote(filename)  # Codifica caracteres especiales

        response['Content-Disposition'] = f'attachment; filename={filename}'
        # Crear el PDF
        p = canvas.Canvas(response, pagesize=letter)
        # Dibujar el logo en la esquina superior derecha
        draw_logo(p, "egatur_logo.png")

        # T√≠tulo
        p.setFont("Helvetica-Bold", 16)
        draw_centered_text(p, "FICHA DE INGRESO DE PERSONAL", 740)
        # Posici√≥n inicial para el contenido
        y_position = 700

        # Secci√≥n I - Datos del Trabajador
        p.setFont("Helvetica-Bold", 12)
        p.drawString(100, y_position, "I. DATOS DEL TRABAJADOR")
        y_position -= 20  # Ajustar la posici√≥n despu√©s del t√≠tulo de la secci√≥n

        y_position = draw_label_value(p, "N¬∫ DE DNI :", persona.dni or "-", y_position)
        y_position = draw_label_value(p, "APELLIDOS Y NOMBRES:", persona.apellidos_nombres or "-", y_position)
        y_position = draw_label_value(p, "FECHA DE NACIMIENTO:", persona.fecha_nacimiento.strftime('%d/%m/%Y') if persona.fecha_nacimiento else "-", y_position)
        y_position = draw_label_value(p, "N¬∫ DE CELULAR:", persona.celular or "-", y_position)
        y_position = draw_label_value(p, "CORREO ELECTR√ìNICO:", persona.correo_personal or "-", y_position)
        y_position = draw_label_value(p, "DIRECCION:", persona.direccion or "-", y_position)
        # L√≠nea de separaci√≥n
        draw_line(p, y_position)
        y_position -= 10

        # Secci√≥n II - Datos Laborales
        p.setFont("Helvetica-Bold", 12)
        p.drawString(100, y_position, "II. DATOS LABORALES")
        y_position -= 20

        y_position = draw_label_value(p, "PERIODO DE INICIO:", persona.periodo_inicio.strftime('%d/%m/%Y') if persona.periodo_inicio else "-", y_position)
        y_position = draw_label_value(p, "PERIODO DE FIN:", persona.periodo_fin.strftime('%d/%m/%Y') if persona.periodo_fin else "-", y_position)
        y_position = draw_label_value(p, "TIPO DE TRABAJADOR:", persona.tipo_trabajador or "-", y_position)
        y_position = draw_label_value(p, "TIPO DE CONTRATO:", persona.tipo_contrato or "-", y_position)
        y_position = draw_label_value(p, "TIPO DE PAGO:", persona.tipo_pago or "-", y_position)
        y_position = draw_label_value(p, "PERIODICIDAD DE INGRESO:", "MENSUAL", y_position)
        y_position = draw_label_value(p, "REMUNERACI√ìN:", f"S/. {persona.remuneracion:.2f}" if persona.remuneracion else "S/. 0.00", y_position)
        y_position = draw_label_value(p, "ASIGNACI√ìN FAMILIAR:", "S√≠" if persona.asignacion_familiar else "No", y_position)

        # L√≠nea de separaci√≥n
        draw_line(p, y_position)
        y_position -= 10

        # Secci√≥n III - Datos de Seguridad Social
        p.setFont("Helvetica-Bold", 12)
        p.drawString(100, y_position, "III. DATOS DE SEGURIDAD SOCIAL")
        y_position -= 20

        y_position = draw_label_value(p, "R√âGIMEN DE SALUD:", (persona.regimen_salud or "-").upper(), y_position)
        p.setFont("Helvetica-Bold", 10)
        p.drawString(100, y_position, "R√âGIMEN PENSIONARIO:")
        p.setFont("Helvetica", 10)
        p.drawString(280, y_position, (persona.regimen_pensionario or "-").upper())
        p.setFont("Helvetica-Bold", 10)
        p.drawString(400, y_position, "DETALLES:")
        p.setFont("Helvetica", 10)
        p.drawString(460, y_position, persona.regimen_pensionario_details or "-")
        draw_line(p, y_position-13)
        y_position -= 20

        # Secci√≥n IV - Datos de la Situaci√≥n Educativa
        p.setFont("Helvetica-Bold", 12)
        p.drawString(100, y_position, "IV. DATOS DE LA SITUACI√ìN EDUCATIVA")
        y_position -= 20

        y_position = draw_label_value(p, "SITUACI√ìN EDUCATIVA:", persona.situacion_educativa or "-", y_position)
        y_position = draw_label_value(p, "TIPO DE INSTRUCCI√ìN:", persona.tipo_instruccion or "-", y_position)
        y_position = draw_label_value(p, "INSTITUCI√ìN:", persona.institucion or "-", y_position)
        y_position = draw_label_value(p, "CARRERA DE ESTUDIO:", persona.carrera_estudio or "-", y_position)
        y_position = draw_label_value(p, "A√ëO DE EGRESO:", persona.ano_egreso or "-", y_position)
        # L√≠nea de separaci√≥n
        draw_line(p, y_position)
        y_position -= 10
        # Secci√≥n V - Horario de Trabajo
        p.setFont("Helvetica-Bold", 12)
        p.drawString(100, y_position, "V. HORARIO DE TRABAJO")
        y_position -= 20

        y_position = draw_label_value(p, "TURNO MA√ëANA:", f"{persona.turno_manana_inicio} - {persona.turno_manana_fin}" if persona.turno_manana_inicio and persona.turno_manana_fin else "-", y_position)
        y_position = draw_label_value(p, "TURNO TARDE:", f"{persona.turno_tarde_inicio} - {persona.turno_tarde_fin}" if persona.turno_tarde_inicio and persona.turno_tarde_fin else "-", y_position)

        # Agregar l√≠neas para firma al final del PDF
        y_position -= 40
        p.line(100, y_position, 300, y_position)
        p.line(350, y_position, 550, y_position)
        p.setFont("Helvetica", 10)
        p.drawString(380, y_position - 15, persona.apellidos_nombres)
        p.showPage()
        p.save()

        return response
    except Exception as e:
        messages.error(request, f"Error al generar el PDF: {str(e)}")
        return redirect("ver_personal")



def draw_label_value(p, label, value, y_position):
    p.setFont("Helvetica-Bold", 10)
    p.drawString(100, y_position, label)  # Escribir la etiqueta en su posici√≥n

    p.setFont("Helvetica", 10)
    x_position = 280  # Aumentamos la posici√≥n X para dar m√°s espacio a la etiqueta
    p.drawString(x_position, y_position, value)  # Escribir el valor m√°s a la derecha

    y_position -= 20  # Bajar la posici√≥n para la siguiente l√≠nea

    return y_position


def draw_line(p, y_position):
    p.setStrokeColor(colors.black)
    p.setLineWidth(1)
    p.line(100, y_position +8, 500, y_position + 8)

def guardar_datos(request):
    if request.method == "POST":

        return redirect("ver_personal")

    return render(request, "ficha_ingreso.html")



def guardar_datos_editados(request, id_personal):
    if request.method == "POST":
        def get_value(field):
            """Devuelve un string vac√≠o ("") si el campo est√° vac√≠o."""
            value = request.POST.get(field, "").strip()
            return value if value else ""

        def get_date(field):
            """Convierte la fecha a formato correcto o devuelve None si est√° vac√≠a."""
            date_value = request.POST.get(field, "").strip()
            try:
                return datetime.strptime(date_value, "%Y-%m-%d").date() if date_value else None
            except ValueError:
                return None  # Evita errores si la fecha no es v√°lida

        def get_time(field):
            """Convierte la hora a formato correcto o devuelve None si est√° vac√≠a."""
            time_value = request.POST.get(field, "").strip()
            try:
                return datetime.strptime(time_value, "%H:%M").time() if time_value else None
            except ValueError:
                return None  # Evita errores si la hora no es v√°lida

        try:
            with transaction.atomic():  # Garantiza que todo se guarde o nada se guarde en caso de error
                ficha = get_object_or_404(Personal, id=id_personal)

                # Obtener los datos del formulario
                banco_id = request.POST.get('nombre_cuenta')
                banco = Banco.objects.get(id=banco_id) if banco_id else None
                # Obtener la sede seleccionada
                sede_id = request.POST.get('sede')
                sede = Local.objects.get(id=sede_id) if sede_id else None
                nuevo_cci = get_value('cci')

                # Buscar la cuenta bancaria anterior con el CCI registrado en la ficha
                cuenta_bancaria = CuentaBancaria.objects.filter(cci=ficha.cci).first()

                if cuenta_bancaria:
                    # Si la cuenta bancaria ya existe, actualizar los datos
                    cuenta_bancaria.nombre_banco = banco.nombre if banco else ""
                    cuenta_bancaria.numero_cuenta = get_value('numero_cuenta')
                    cuenta_bancaria.cci = nuevo_cci
                    cuenta_bancaria.save()
                else:
                    # Si no existe, crear una nueva cuenta bancaria
                    cuenta_bancaria = CuentaBancaria.objects.create(
                        proveedor=None,  # Ajustar si es necesario asignar un proveedor
                        nombre_banco=banco.nombre if banco else "",
                        numero_cuenta=get_value('numero_cuenta'),
                        cci=nuevo_cci
                    )

                # Actualizar la ficha del personal con los nuevos valores
                ficha.dni = get_value("dni")
                ficha.apellidos_nombres = get_value("apellidos_nombres")
                ficha.fecha_nacimiento = get_date("fecha_nacimiento")
                ficha.celular = get_value("celular")
                ficha.correo_personal = get_value("correo_personal")
                ficha.correo_corporativo = get_value("correo_corporativo")
                ficha.direccion = get_value("direccion")
                ficha.periodo_inicio = get_date("periodo_inicio")
                ficha.periodo_fin = get_date("periodo_fin")
                ficha.tipo_trabajador = get_value("tipo_trabajador")
                ficha.tipo_contrato = get_value("tipo_contrato")
                ficha.tipo_pago = get_value("tipo_pago")
                ficha.asignacion_familiar = bool(request.POST.get("asignacion_familiar"))
                ficha.regimen_salud = get_value("regimen_salud")
                ficha.regimen_pensionario = get_value("regimen_pensionario")
                ficha.regimen_pensionario_details = get_value("regimen_pensionario_details")
                ficha.situacion_educativa = get_value("situacion_educativa")
                ficha.tipo_instruccion = get_value("tipo_instruccion")
                ficha.institucion = get_value("institucion")
                ficha.carrera_estudio = get_value("carrera_estudio")
                ficha.ocupacion = get_value("ocupacion")
                ficha.remuneracion = get_value("remuneracion")
                ficha.ano_egreso = get_value("ano_egreso")
                ficha.nombre_cuenta = banco.nombre if banco else ""
                ficha.numero_cuenta = get_value('numero_cuenta')
                ficha.cci = nuevo_cci
                ficha.turno_manana_inicio = get_time("turno_manana_inicio")
                ficha.turno_manana_fin = get_time("turno_manana_fin")
                ficha.turno_tarde_inicio = get_time("turno_tarde_inicio")
                ficha.turno_tarde_fin = get_time("turno_tarde_fin")
                ficha.observacion = get_value("observaciones")
                ficha.local = sede

                ficha.save()

                messages.success(request, "Ficha  actualizada")
                return redirect("ver_personal")

        except Exception as e:
            messages.error(request, f"Error al actualizar la ficha: {str(e)}")
            return redirect("ver_personal")


@csrf_exempt
def eliminar_concepto(request, id):
    if request.method == 'POST':
        try:
            concepto = Concepto.objects.get(id=id)
            concepto.delete()
            return JsonResponse({'status': 'success'})
        except Concepto.DoesNotExist:
            return JsonResponse({'error': 'El concepto no existe.'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def crear_concepto(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        nivel = int(request.POST.get('nivel'))
        id_padre = request.POST.get('id_padre')

        if nivel != 1 and not id_padre:
            return JsonResponse({'error': 'El nivel 2 y 3 requieren un concepto padre.'}, status=400)

        id_concepto_padre = Concepto.objects.get(id=id_padre) if id_padre else None
        Concepto.objects.create(
            concepto_nombre=nombre,
            nivel=nivel,
            id_concepto_padre=id_concepto_padre,
        )
        return JsonResponse({'status': 'success'})


@csrf_exempt
def editar_concepto(request, id):
    if request.method == 'POST':
        concepto = Concepto.objects.get(id=id)
        concepto.concepto_nombre = request.POST.get('nombre')
        concepto.save()
        return JsonResponse({'status': 'success'})


def build_concepto_hierarchy(conceptos):
    """Crea una jerarqu√≠a de conceptos basada en sus niveles."""
    hierarchy = []

    # Agrupar conceptos por nivel
    nivel1 = conceptos.filter(nivel=1)
    nivel2 = conceptos.filter(nivel=2)
    nivel3 = conceptos.filter(nivel=3)

    for nivel1_item in nivel1:
        nivel1_dict = {
            "concepto": nivel1_item,
            "subconceptos": []
        }
        for nivel2_item in nivel2.filter(id_concepto_padre=nivel1_item):
            nivel2_dict = {
                "concepto": nivel2_item,
                "subconceptos": []
            }
            for nivel3_item in nivel3.filter(id_concepto_padre=nivel2_item):
                nivel2_dict["subconceptos"].append({
                    "concepto": nivel3_item,
                    "subconceptos": []
                })
            nivel1_dict["subconceptos"].append(nivel2_dict)
        hierarchy.append(nivel1_dict)

    return hierarchy

def conceptos(request):
    conceptos = Concepto.objects.all()
    conceptos_hierarchy = build_concepto_hierarchy(conceptos)

    return render(request, 'conceptos.html', {
        'conceptos_hierarchy': conceptos_hierarchy
    })
def reportes(request):
    # Filtrar los conceptos de nivel 1 y nivel 2
    conceptos_nivel_1 = Concepto.objects.filter(nivel=1)
    conceptos_nivel_2 = Concepto.objects.filter(nivel=2)
    proveedores = Proveedor.objects.all()  # Aqu√≠ estaba el error en "Proveedor.all()"
    # Pasa todos los conjuntos de datos al contexto
    return render(request, 'reportes.html', {
        'conceptos_nivel_1': conceptos_nivel_1,
        'conceptos_nivel_2': conceptos_nivel_2,
        'proveedores': proveedores,  # Ahora tambi√©n pasamos proveedores
    })

import io
from datetime import date
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from myapp.models import Proveedor, Gasto  # Ajusta seg√∫n tu app

def trigger_error(request):
    # Esto arrojar√° un error ZeroDivisionError, provocando un error 500.
    division_by_zero = 1 / 0

def custom_404(request, exception):
    return render(request, '404.html', status=404)

def custom_500(request):
    return render(request, '500.html', status=500)
def reporte_proveedor_pdf(request):
    # Obtener el id del proveedor desde el par√°metro GET (ejemplo: ?proveedor=120)
    proveedor_id = request.GET.get('proveedor')
    if not proveedor_id:
        return HttpResponse("Proveedor no especificado.", status=400)

    try:
        proveedor = Proveedor.objects.get(pk=proveedor_id)
    except Proveedor.DoesNotExist:
        return HttpResponse("Proveedor no encontrado.", status=404)

    # Mes actual
    current_month = date.today().month

    # Obtener registros de Gasto y Rendicion que correspondan al proveedor y al mes actual
    gastos_queryset = Gasto.objects.filter(nombre_proveedor=proveedor, fecha_gasto__month=current_month).exclude(tipo_comprobante="Requerimiento",gasto_origen__isnull=True)
    rendiciones_queryset = Rendicion.objects.filter(proveedor=proveedor, fecha_operacion__month=current_month)

    # Combinar ambos querysets en una lista
    registros = list(gastos_queryset) + list(rendiciones_queryset)

    # Funci√≥n auxiliar para obtener la fecha (diferente en cada modelo)
    def get_fecha(registro):
        if hasattr(registro, 'fecha_gasto') and registro.fecha_gasto:
            return registro.fecha_gasto
        elif hasattr(registro, 'fecha_operacion') and registro.fecha_operacion:
            return registro.fecha_operacion
        return date.today()

    # Ordenar los registros por fecha
    registros.sort(key=get_fecha)

    # Crear el buffer y documento PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Estilo para la columna Concepto (para que el texto se ajuste)
    concept_style = ParagraphStyle(
        name='ConceptStyle',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        leading=12,
        alignment=0  # Alineaci√≥n izquierda
    )

    # T√≠tulo del reporte
    titulo = f"Reporte de Gastos y Rendiciones para el Proveedor: {proveedor.razon_social} - {date.today().strftime('%B %Y')}"
    elements.append(Paragraph(titulo, styles['Title']))
    elements.append(Spacer(1, 12))

    # Preparar los datos de la tabla
    data = []
    headers = ["Fecha", "Concepto", "Tipo Comprobante", "Importe"]
    data.append(headers)
    total_importe = 0

    for reg in registros:
        # Obtener la fecha
        if hasattr(reg, 'fecha_gasto') and reg.fecha_gasto:
            fecha_str = reg.fecha_gasto.strftime("%Y-%m-%d")
        elif hasattr(reg, 'fecha_operacion') and reg.fecha_operacion:
            fecha_str = reg.fecha_operacion.strftime("%Y-%m-%d")
        else:
            fecha_str = ""

        # Determinar el concepto de mayor nivel (nivel 3, sino nivel 2, sino nivel 1)
        if hasattr(reg, 'concepto_nivel_3') and reg.concepto_nivel_3:
            concepto_str = reg.concepto_nivel_3.concepto_nombre
        elif hasattr(reg, 'concepto_nivel_2') and reg.concepto_nivel_2:
            concepto_str = reg.concepto_nivel_2.concepto_nombre
        elif hasattr(reg, 'concepto_nivel_1') and reg.concepto_nivel_1:
            concepto_str = reg.concepto_nivel_1.concepto_nombre
        else:
            concepto_str = ""

        # Tipo de comprobante (ambos modelos tienen 'tipo_comprobante')
        tipo_comp = reg.tipo_comprobante if hasattr(reg, 'tipo_comprobante') and reg.tipo_comprobante else ""

        # Importe
        importe_val = float(reg.importe) if reg.importe else 0
        total_importe += importe_val

        # Envolver el concepto en un Paragraph para que se ajuste en la celda
        concepto_paragraph = Paragraph(concepto_str, concept_style)

        data.append([fecha_str, concepto_paragraph, tipo_comp, f"{importe_val:,.2f}"])

    # Agregar la fila de Total
    data.append(["", "", "Total", f"{total_importe:,.2f}"])

    # Crear la tabla con anchos de columna fijos
    table = Table(data, colWidths=[100, 250, 150, 100])
    table_style = TableStyle([
        # Encabezado: fondo celeste y texto blanco
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        # Fila de Total: fondo gris y negrita
        ('BACKGROUND', (0, -1), (-1, -1), colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        # Bordes para toda la tabla
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(table_style)
    elements.append(table)

    # Construir el PDF
    doc.build(elements)
    buffer.seek(0)

    # Preparar la respuesta HTTP para la descarga
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"reporte_proveedor_{proveedor.razon_social}_{date.today().strftime('%B')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



def reporte_proveedor_excel(request):
    # Obtener el id del proveedor desde el par√°metro GET (por ejemplo, ?proveedor=120)
    proveedor_id = request.GET.get('proveedor')
    if not proveedor_id:
        return HttpResponse("Proveedor no especificado.", status=400)

    try:
        proveedor = Proveedor.objects.get(pk=proveedor_id)
    except Proveedor.DoesNotExist:
        return HttpResponse("Proveedor no encontrado.", status=404)

    # Obtener el mes actual (para filtrar los gastos del mes actual)
    current_month = date.today().month

    # Filtrar los gastos usando el campo 'nombre_proveedor'
    gastos = Gasto.objects.filter(nombre_proveedor=proveedor, fecha_gasto__month=current_month)

    # Crear el libro y la hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Proveedor"

    # T√≠tulo: crear una fila superior con celdas fusionadas
    titulo = f"Reporte de Gastos para el Proveedor: {proveedor.razon_social} - {date.today().strftime('%B %Y')}"
    ws.merge_cells('A1:F1')
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Encabezados de la tabla (fila 3)
    headers = ["Fecha Gasto", "Concepto", "Local", "Tipo Comprobante", "Importe", "Observaci√≥n"]
    header_row = 3
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Rellenar la tabla con los datos (a partir de la fila 4)
    data_start_row = header_row + 1
    current_row = data_start_row
    total_importe = 0

    for gasto in gastos:
        # Determinar el concepto de mayor nivel para el gasto.
        if gasto.concepto_nivel_3:
            concepto_str = gasto.concepto_nivel_3.concepto_nombre
        elif gasto.concepto_nivel_2:
            concepto_str = gasto.concepto_nivel_2.concepto_nombre
        elif gasto.concepto_nivel_1:
            concepto_str = gasto.concepto_nivel_1.concepto_nombre
        else:
            concepto_str = ""

        # Formatear la fecha (si est√° disponible)
        fecha_str = gasto.fecha_gasto.strftime("%Y-%m-%d") if gasto.fecha_gasto else ""
        # Convertir el campo local a cadena (utiliza __str__ del modelo Local)
        local_str = str(gasto.local)
        tipo_comp = gasto.tipo_comprobante
        importe_val = float(gasto.importe)
        observacion_val = gasto.observacion

        ws.cell(row=current_row, column=1, value=fecha_str)
        ws.cell(row=current_row, column=2, value=concepto_str)
        ws.cell(row=current_row, column=3, value=local_str)
        ws.cell(row=current_row, column=4, value=tipo_comp)
        imp_cell = ws.cell(row=current_row, column=5, value=importe_val)
        imp_cell.number_format = '#,##0.00'
        ws.cell(row=current_row, column=6, value=observacion_val)

        total_importe += importe_val
        current_row += 1

    # Agregar fila de total (despu√©s de los datos)
    total_row = current_row + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    total_label = ws.cell(row=total_row, column=1, value="Total")
    total_label.font = Font(bold=True)
    total_label.alignment = Alignment(horizontal='right')
    total_cell = ws.cell(row=total_row, column=5, value=total_importe)
    total_cell.font = Font(bold=True)
    total_cell.fill = header_fill
    total_cell.number_format = '#,##0.00'

    # Ajustar el ancho de las columnas seg√∫n el contenido usando get_column_letter
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(i)
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[column].width = max_length + 2

    # Preparar la respuesta HTTP con el Excel para descarga directa
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"reporte_proveedor_{proveedor.razon_social}_{date.today().strftime('%B')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
def convertir_a_float(valor):
    """Convierte un string con coma o punto decimal a float correctamente."""
    if isinstance(valor, str):
        valor = valor.replace(',', '.')  # Reemplazar coma por punto
    try:
        return float(valor)
    except ValueError:
        return 0.00  # Si hay error, devolver 0.00


def convertir_a_float(valor):
    try:
        if isinstance(valor, str):
            return float(valor.replace(',', '.'))
        return float(valor)
    except (ValueError, TypeError):
        return 0.00

@login_required
def generar_reporte_diario(request):
    try:
        if not request.user.is_authenticated:
            messages.error(request, "Usuario no autenticado")
            return HttpResponse("Error: Usuario no autenticado", status=403)

        hoy = date.today()

        # Configuraci√≥n de estilos
        HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        SECTION_FILL = PatternFill(start_color="D6E1F3", end_color="D6E1F3", fill_type="solid")
        TOTAL_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        CLOSING_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        BOLD_FONT = Font(bold=True, color="000000")
        HEADER_FONT = Font(bold=True, color="FFFFFF")
        MONEY_FORMAT = '#,##0.00'
        CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
        RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
        BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Diario"

        # Encabezado principal
        ws.merge_cells('A1:D1')
        header_cell = ws['A1']
        header_cell.value = "REPORTE DIARIO"
        header_cell.fill = HEADER_FILL
        header_cell.font = HEADER_FONT
        header_cell.alignment = CENTER_ALIGN

        # Fecha del reporte
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Fecha: {hoy.strftime('%d de %B de %Y')}"
        ws['A2'].alignment = CENTER_ALIGN

        # I. SALDO DE APERTURA
        saldo_inicial = convertir_a_float(request.GET.get('saldo_inicial', '0.00'))
        ws.append([])
        row_num = ws.max_row + 1
        ws.merge_cells(f'A{row_num}:C{row_num}')
        ws[f'A{row_num}'] = "I. SALDO DE APERTURA DE CAJA"
        ws[f'A{row_num}'].fill = SECTION_FILL
        ws[f'A{row_num}'].font = BOLD_FONT
        ws[f'D{row_num}'] = saldo_inicial
        ws[f'D{row_num}'].number_format = MONEY_FORMAT

        # II. RECEPCI√ìN DE EFECTIVO
        ingresos = Ingreso.objects.filter(usuario_creador=request.user, fecha_ingreso=hoy)
        total_ingresos = sum(convertir_a_float(i.importe) for i in ingresos)

        ws.append([])
        row_num = ws.max_row + 1
        ws.merge_cells(f'A{row_num}:D{row_num}')
        ws[f'A{row_num}'] = "II. RECEPCI√ìN DE EFECTIVO"
        ws[f'A{row_num}'].fill = SECTION_FILL
        ws[f'A{row_num}'].font = BOLD_FONT

        # Encabezados de tabla
        headers = ["ID", "Fecha", "Comentario", "Monto (S/.)"]
        ws.append(headers)
        for col in range(1, 5):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN

        # Datos de ingresos
        for ingreso in ingresos:
            ws.append([
                ingreso.id,
                ingreso.fecha_ingreso.strftime('%d/%m/%Y') if ingreso.fecha_ingreso else '',
                ingreso.observacion or "",
                convertir_a_float(ingreso.importe)
            ])
            ws.cell(row=ws.max_row, column=4).number_format = MONEY_FORMAT

        # Total ingresos
        ws.append(["TOTAL INGRESOS", "", "", total_ingresos])
        for col in range(1, 5):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = TOTAL_FILL
            cell.font = BOLD_FONT

        # III. GASTOS
        gastos = Gasto.objects.filter(usuario_creador=request.user, fecha_gasto=hoy)
        total_gastos = sum(convertir_a_float(g.importe) for g in gastos)

        ws.append([])
        row_num = ws.max_row + 1
        ws.merge_cells(f'A{row_num}:C{row_num}')
        ws[f'A{row_num}'] = "III. GASTOS"
        ws[f'A{row_num}'].fill = SECTION_FILL
        ws[f'A{row_num}'].font = BOLD_FONT
        ws[f'D{row_num}'] = total_gastos
        ws[f'D{row_num}'].number_format = MONEY_FORMAT

        # IV. SALDO AL CIERRE
        saldo_cierre = saldo_inicial + total_ingresos - total_gastos
        ws.append([])
        row_num = ws.max_row + 1
        ws.merge_cells(f'A{row_num}:C{row_num}')
        ws[f'A{row_num}'] = "IV. SALDO AL CIERRE DEL D√çA"
        ws[f'A{row_num}'].fill = CLOSING_FILL
        ws[f'A{row_num}'].font = BOLD_FONT
        ws[f'D{row_num}'] = saldo_cierre
        ws[f'D{row_num}'].number_format = MONEY_FORMAT

        # Ajustar formato de celdas
        def is_main_merged(cell):
            for range in ws.merged_cells.ranges:
                if cell.coordinate in range:
                    return cell.coordinate == range.start_cell
            return True

        for row in ws.iter_rows():
            for cell in row:
                if is_main_merged(cell):
                    cell.border = BORDER
                    if cell.column == 4:
                        cell.alignment = RIGHT_ALIGN

        # Ajustar anchos de columnas
        column_widths = {'A': 10, 'B': 15, 'C': 40, 'D': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Generar respuesta
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="Reporte_Diario_{hoy.strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response

    except Exception as e:
        messages.error(request, f"Error generando reporte: {str(e)}")
        return HttpResponse(f"Error generando reporte: {str(e)}", status=500)
def reporte_mensual(request):
    if request.method == "POST":
        mes_inicio_str = request.POST.get("mes_inicio")
        mes_fin_str = request.POST.get("mes_fin")
        if not mes_inicio_str or not mes_fin_str:
            return HttpResponse("Debe seleccionar el mes de inicio y el mes de fin.", status=400)

        try:
            mes_inicio = int(mes_inicio_str)
            mes_fin = int(mes_fin_str)
        except ValueError:
            return HttpResponse("Mes inv√°lido.", status=400)

        if mes_inicio > mes_fin:
            return HttpResponse("El mes de inicio no puede ser mayor que el mes de fin.", status=400)

        # Crear libro y hoja de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Reporte Mensual {mes_inicio:02d} a {mes_fin:02d}"

        # Estilos
        bold_font = Font(bold=True)
        bold_gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        bold_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        italic_bold_font = Font(bold=True, italic=True)
        gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Nombres de los meses (lista de 12 elementos)
        nombres_meses = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]
        # Seleccionar los nombres de los meses que est√°n en el rango (√≠ndices de 0 a 11)
        meses_rango = nombres_meses[mes_inicio-1:mes_fin]

        # Encabezado de la hoja: Concepto, columnas para cada mes en el rango y Total
        ws.append(['Concepto'] + meses_rango + ['Total'])
        for cell in ws[1]:
            cell.font = bold_font

        # Variables para acumulaci√≥n de totales
        total_general_mes = [0] * (mes_fin - mes_inicio + 1)
        total_general = 0

        # Procesar conceptos de nivel 1
        conceptos_nivel_1 = Concepto.objects.filter(nivel=1)
        for concepto1 in conceptos_nivel_1:
            valores_mensuales = []
            total_concepto = 0

            # Iterar s√≥lo sobre los meses del rango
            for idx, mes_numero in enumerate(range(mes_inicio, mes_fin+1)):
                total_mes = (
                    Gasto.objects.filter(
                        fecha_gasto__month=mes_numero,
                        concepto_nivel_1=concepto1
                    ).aggregate(total=Sum('importe'))['total'] or 0
                ) + (
                    Rendicion.objects.filter(
                        fecha_operacion__month=mes_numero,
                        concepto_nivel_1=concepto1
                    ).aggregate(total=Sum('importe'))['total'] or 0
                )

                # Para "GASTOS DIVERSOS", sumar tambi√©n requerimientos no rendidos (excluyendo aquellos ya vinculados a rendiciones)
                if concepto1.concepto_nombre == "GASTOS DIVERSOS":
                    total_mes += (
                        Gasto.objects.filter(
                            fecha_gasto__month=mes_numero,
                            tipo_comprobante="Requerimiento",
                            rendido=False,
                            rendiciones_gasto__isnull=True
                        ).aggregate(total=Sum('importe'))['total'] or 0
                    )

                valores_mensuales.append(total_mes)
                total_concepto += total_mes
                total_general_mes[idx] += total_mes

            total_general += total_concepto
            ws.append([concepto1.concepto_nombre] + valores_mensuales + [total_concepto])
            # Aplicar estilos
            for cell in ws[ws.max_row]:
                cell.font = bold_font
            ws.cell(row=ws.max_row, column=1).fill = bold_gray_fill

            # Procesar conceptos de nivel 2
            conceptos_nivel_2 = Concepto.objects.filter(id_concepto_padre=concepto1)
            for concepto2 in conceptos_nivel_2:
                valores_mensuales = []
                total_concepto = 0
                for mes_numero in range(mes_inicio, mes_fin+1):
                    if concepto2.concepto_nombre == "ENTREGAS DE EFECTIVO PENDIENTES POR RENDIR CUENTAS":
                        total_mes = (
                            Gasto.objects.filter(
                                fecha_gasto__month=mes_numero,
                                tipo_comprobante="Requerimiento",
                                rendido=False,
                                gasto_origen__isnull=True  # Asegura que gasto_origen est√© vac√≠o
                            ).aggregate(total=Sum('importe'))['total'] or 0
                        )
                    else:
                        total_mes = (
                            Gasto.objects.filter(
                                fecha_gasto__month=mes_numero,
                                concepto_nivel_2=concepto2
                            ).aggregate(total=Sum('importe'))['total'] or 0
                        ) + (
                            Rendicion.objects.filter(
                                fecha_operacion__month=mes_numero,
                                concepto_nivel_2=concepto2
                            ).aggregate(total=Sum('importe'))['total'] or 0
                        )
                    valores_mensuales.append(total_mes)
                    total_concepto += total_mes
                ws.append(["   " + concepto2.concepto_nombre] + valores_mensuales + [total_concepto])
                ws.cell(row=ws.max_row, column=1).font = bold_font

                # Procesar conceptos de nivel 3
                conceptos_nivel_3 = Concepto.objects.filter(id_concepto_padre=concepto2)
                for concepto3 in conceptos_nivel_3:
                    valores_mensuales = []
                    total_concepto = 0
                    for mes_numero in range(mes_inicio, mes_fin+1):
                        total_mes = (
                            Gasto.objects.filter(
                                fecha_gasto__month=mes_numero,
                                concepto_nivel_3=concepto3
                            ).aggregate(total=Sum('importe'))['total'] or 0
                        ) + (
                            Rendicion.objects.filter(
                                fecha_operacion__month=mes_numero,
                                concepto_nivel_3=concepto3
                            ).aggregate(total=Sum('importe'))['total'] or 0
                        )
                        valores_mensuales.append(total_mes)
                        total_concepto += total_mes
                    ws.append(["      " + concepto3.concepto_nombre] + valores_mensuales + [total_concepto])
                    ws.cell(row=ws.max_row, column=1).fill = yellow_fill

        # Agregar fila final con totales por mes y total general
        ws.append(["Total por Mes"] + total_general_mes + [total_general])
        for cell in ws[ws.max_row]:
            cell.font = bold_font
            cell.fill = bold_blue_fill

        # Generar respuesta HTTP para la descarga del archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Mensual_{mes_inicio:02d}_{mes_fin:02d}_{date.today().year}.xlsx"'
        wb.save(response)
        return response


def reporte_anual(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Anual"

    # Estilos
    bold_font = Font(bold=True)
    bold_gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    bold_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Nombres de los meses
    meses = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]

    # Encabezado
    ws.append(['Concepto'] + meses + ['Total Anual'])
    for cell in ws[1]:
        cell.font = bold_font

    # Variables para el total final
    total_general_mes = [0] * 12
    total_general_anual = 0

    # Conceptos de nivel 1
    conceptos_nivel_1 = Concepto.objects.filter(nivel=1)
    for concepto1 in conceptos_nivel_1:
        valores_mensuales = []
        total_anual = 0

        for mes_numero in range(1, 13):
            total_mes_mes = (
                Gasto.objects.filter(fecha_gasto__month=mes_numero, concepto_nivel_1=concepto1)
                .aggregate(total=Sum('importe'))['total'] or 0
            ) + (
                Rendicion.objects.filter(fecha_operacion__month=mes_numero, concepto_nivel_1=concepto1)
                .aggregate(total=Sum('importe'))['total'] or 0
            )

            # Si el concepto es "GASTOS DIVERSOS", sumar tambi√©n los Requerimientos no rendidos
            if concepto1.concepto_nombre == "GASTOS DIVERSOS":
                total_mes_mes += (
                    Gasto.objects.filter(fecha_gasto__month=mes_numero, tipo_comprobante="Requerimiento", rendido=False)
                    .aggregate(total=Sum('importe'))['total'] or 0
                )

            valores_mensuales.append(total_mes_mes)
            total_anual += total_mes_mes
            total_general_mes[mes_numero - 1] += total_mes_mes

        total_general_anual += total_anual
        ws.append([concepto1.concepto_nombre] + valores_mensuales + [total_anual])
        for cell in ws[ws.max_row]:
            cell.font = bold_font
        ws[ws.max_row][0].fill = bold_gray_fill

        # Conceptos de nivel 2
        conceptos_nivel_2 = Concepto.objects.filter(id_concepto_padre=concepto1)
        for concepto2 in conceptos_nivel_2:
            valores_mensuales = []
            total_anual = 0

            for mes_numero in range(1, 13):
                total_mes_mes = (
                    Gasto.objects.filter(fecha_gasto__month=mes_numero, concepto_nivel_2=concepto2)
                    .aggregate(total=Sum('importe'))['total'] or 0
                ) + (
                    Rendicion.objects.filter(fecha_operacion__month=mes_numero, concepto_nivel_2=concepto2)
                    .aggregate(total=Sum('importe'))['total'] or 0
                )

                # Si es "ENTREGAS DE EFECTIVO PENDIENTES POR RENDIR CUENTAS", sumar Requerimientos no rendidos
                if concepto2.concepto_nombre == "ENTREGAS DE EFECTIVO PENDIENTES POR RENDIR CUENTAS":
                    total_mes_mes = (
                        Gasto.objects.filter(fecha_gasto__month=mes_numero, tipo_comprobante="Requerimiento", rendido=False,gasto_origen__isnull=True)
                        .aggregate(total=Sum('importe'))['total'] or 0
                    )

                valores_mensuales.append(total_mes_mes)
                total_anual += total_mes_mes

            ws.append(["   " + concepto2.concepto_nombre] + valores_mensuales + [total_anual])
            ws[ws.max_row][0].font = bold_font

    # Agregar la fila final con el total por mes y total anual
    ws.append(["Total por Mes"] + total_general_mes + [total_general_anual])
    for cell in ws[ws.max_row]:
        cell.font = bold_font
        cell.fill = bold_blue_fill

    # Generar respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Anual_{date.today().year}.xlsx"'
    wb.save(response)
    return response




@login_required
def ver_personal(request):
    today = date.today()  # Fecha actual
    hace_7_dias = today + timedelta(days=7)  # Hace una semana

    personal = Personal.objects.all()  # Obtener los registros de Personal
    locales = Local.objects.all()

    return render(request, 'ver_personal.html', {
        'personal': personal,
        'today': today,
        'locales':locales,
        'hace_7_dias': hace_7_dias  # Pasamos la fecha al template
    })

def reporte_diario_conceptos(request):
    # Obtener fechas del formulario
    fecha_inicio = request.POST.get('fecha_inicio')
    fecha_fin = request.POST.get('fecha_fin')

    if not fecha_inicio or not fecha_fin:
        return HttpResponse("Error: Debes proporcionar una fecha de inicio y una fecha de fin.", status=400)

    # Convertir fechas a formato datetime
    fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d")

    # Crear el workbook y la hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Diario Conceptos"

    # Estilos
    bold_font = Font(bold=True)
    bold_gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    bold_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Generar lista de fechas en el rango
    rango_fechas = []
    current_date = fecha_inicio
    while current_date <= fecha_fin:
        rango_fechas.append(current_date.strftime("%Y-%m-%d"))
        current_date += timedelta(days=1)

    # Encabezado con los d√≠as en el rango
    ws.append(['Concepto'] + rango_fechas + ['Total General'])
    for cell in ws[1]:
        cell.font = bold_font

    total_general_dias = [0] * len(rango_fechas)  # Totales por d√≠a
    total_general = 0  # Total global

    # Obtener los conceptos de nivel 1
    conceptos_nivel_1 = Concepto.objects.filter(nivel=1)

    for concepto1 in conceptos_nivel_1:
        valores_diarios = []
        total_concepto1 = 0

        # Iterar por cada d√≠a en el rango
        for index, fecha in enumerate(rango_fechas):
            fecha_dt = datetime.strptime(fecha, "%Y-%m-%d")
            total_dia = (
                Gasto.objects.filter(fecha_gasto=fecha_dt, concepto_nivel_1=concepto1)
                .aggregate(total=Sum('importe'))['total'] or 0
            ) + (
                Rendicion.objects.filter(fecha_operacion=fecha_dt, concepto_nivel_1=concepto1)
                .aggregate(total=Sum('importe'))['total'] or 0
            )

            # Si el concepto es "GASTOS DIVERSOS", sumar requerimientos no rendidos
            if concepto1.concepto_nombre == "GASTOS DIVERSOS":
                total_dia += (
                    Gasto.objects.filter(
                        fecha_gasto=fecha_dt,
                        tipo_comprobante="Requerimiento",
                        rendido=False
                    ).aggregate(total=Sum('importe'))['total'] or 0
                )

            valores_diarios.append(total_dia)
            total_concepto1 += total_dia
            total_general_dias[index] += total_dia

        total_general += total_concepto1
        ws.append([concepto1.concepto_nombre] + valores_diarios + [total_concepto1])
        for cell in ws[ws.max_row]:
            cell.font = bold_font
        ws[ws.max_row][0].fill = bold_gray_fill

        # Conceptos de nivel 2
        conceptos_nivel_2 = Concepto.objects.filter(id_concepto_padre=concepto1)
        for concepto2 in conceptos_nivel_2:
            valores_diarios = []
            total_concepto2 = 0

            for index, fecha in enumerate(rango_fechas):
                fecha_dt = datetime.strptime(fecha, "%Y-%m-%d")
                total_dia = (
                    Gasto.objects.filter(fecha_gasto=fecha_dt, concepto_nivel_2=concepto2)
                    .aggregate(total=Sum('importe'))['total'] or 0
                ) + (
                    Rendicion.objects.filter(fecha_operacion=fecha_dt, concepto_nivel_2=concepto2)
                    .aggregate(total=Sum('importe'))['total'] or 0
                )

                valores_diarios.append(total_dia)
                total_concepto2 += total_dia

            # Omitir conceptos de nivel 2 con solo ceros
            if total_concepto2 > 0:
                ws.append(["   " + concepto2.concepto_nombre] + valores_diarios + [total_concepto2])
                ws[ws.max_row][0].font = bold_font

    # Agregar la fila final con los totales por d√≠a y total general
    ws.append(["Total General"] + total_general_dias + [total_general])
    for cell in ws[ws.max_row]:
        cell.font = bold_font
        cell.fill = bold_blue_fill

    # Generar respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Diario_{fecha_inicio.strftime("%Y-%m-%d")}_al_{fecha_fin.strftime("%Y-%m-%d")}.xlsx"'
    wb.save(response)
    return response

def guardar_datos1(request):
    if request.method == 'POST':
        def get_value(field):
            """Devuelve un string vac√≠o ("") si el campo est√° vac√≠o."""
            value = request.POST.get(field, "").strip()
            return value if value else ""

        def get_date(field):
            """Convierte la fecha a formato correcto o devuelve un JsonResponse si hay error."""
            date_value = request.POST.get(field, "").strip()
            if not date_value:
                return None
            try:
                return datetime.strptime(date_value, "%Y-%m-%d").date()
            except ValueError as e:
                return JsonResponse({'error': f'Fecha inv√°lida en {field}: {str(e)}'}, status=400)

        def get_time(field):
            """Convierte la hora a formato correcto o devuelve un JsonResponse si hay error."""
            time_value = request.POST.get(field, "").strip()
            if not time_value:
                return None
            try:
                return datetime.strptime(time_value, "%H:%M").time()
            except ValueError as e:
                return JsonResponse({'error': f'Hora inv√°lida en {field}: {str(e)}'}, status=400)

        # Verificar si ya existe un proveedor con el mismo DNI
        if Proveedor.objects.filter(ruc_dni=request.POST.get('dni')).exists():
            return JsonResponse({
                'error': 'Ya existe un proveedor con ese DNI. Elimine el proveedor antes de continuar.'
            }, status=400)

        try:
            with transaction.atomic():  # Garantiza que todo se guarde o nada se guarde
                # Obtener el banco seleccionado
                banco_id = request.POST.get('nombre_cuenta')
                try:
                    banco = Banco.objects.get(id=banco_id) if banco_id else None
                except Banco.DoesNotExist as e:
                    return JsonResponse({'error': f'Banco no encontrado: {str(e)}'}, status=404)

                # Obtener la sede (Local)
                local_id = request.POST.get('sede')
                try:
                    local = Local.objects.get(id=local_id) if local_id else None
                except Local.DoesNotExist as e:
                    return JsonResponse({'error': f'Local no encontrado: {str(e)}'}, status=404)

                # Guardar los datos del Personal
                personal = Personal(
                    dni=get_value('dni'),
                    apellidos_nombres=get_value('apellidos_nombres'),
                    fecha_nacimiento=get_date('fecha_nacimiento'),
                    celular=get_value('celular'),
                    correo_personal=get_value('correo_personal'),
                    correo_corporativo=get_value('correo_corporativo'),
                    direccion=get_value('direccion'),
                    periodo_inicio=get_date('periodo_inicio'),
                    periodo_fin=get_date('periodo_fin'),
                    tipo_trabajador=get_value('tipo_trabajador'),
                    tipo_contrato=get_value('tipo_contrato'),
                    tipo_pago=get_value('tipo_pago'),
                    nombre_cuenta=banco.nombre if banco else "",
                    numero_cuenta=get_value('numero_cuenta'),
                    asignacion_familiar=request.POST.get('asignacion_familiar') == 'on',
                    ocupacion=get_value('ocupacion'),
                    remuneracion=get_value('remuneracion'),
                    regimen_salud=get_value('regimen_salud'),
                    regimen_pensionario=get_value('regimen_pensionario'),
                    situacion_educativa=get_value('situacion_educativa'),
                    tipo_instruccion=get_value('tipo_instruccion'),
                    institucion=get_value('institucion'),
                    carrera_estudio=get_value('carrera_estudio'),
                    ano_egreso=get_value('ano_egreso'),
                    cci=get_value('cci'),
                    # Guardar los horarios de trabajo
                    turno_manana_inicio=get_time('turno_manana_inicio'),
                    turno_manana_fin=get_time('turno_manana_fin'),
                    turno_tarde_inicio=get_time('turno_tarde_inicio'),
                    turno_tarde_fin=get_time('turno_tarde_fin'),
                    local=local
                )

                if isinstance(personal.fecha_nacimiento, JsonResponse):
                    return personal.fecha_nacimiento
                if isinstance(personal.periodo_inicio, JsonResponse):
                    return personal.periodo_inicio
                if isinstance(personal.periodo_fin, JsonResponse):
                    return personal.periodo_fin
                if isinstance(personal.turno_manana_inicio, JsonResponse):
                    return personal.turno_manana_inicio
                if isinstance(personal.turno_manana_fin, JsonResponse):
                    return personal.turno_manana_fin
                if isinstance(personal.turno_tarde_inicio, JsonResponse):
                    return personal.turno_tarde_inicio
                if isinstance(personal.turno_tarde_fin, JsonResponse):
                    return personal.turno_tarde_fin

                personal.save()

                # Crear un Proveedor con los datos del Personal
                proveedor = Proveedor(
                    ruc_dni=personal.dni,
                    razon_social=personal.apellidos_nombres,
                    telefono=personal.celular,
                    nombre_contacto=""
                )
                proveedor.save()

                # Crear una cuenta bancaria asociada al proveedor
                cuenta_bancaria = CuentaBancaria(
                    proveedor=proveedor,
                    nombre_banco=personal.nombre_cuenta,
                    numero_cuenta=personal.numero_cuenta,
                    cci=personal.cci
                )
                cuenta_bancaria.save()

            return JsonResponse({
                'success': 'Datos guardados correctamente y proveedor creado.'
            }, status=200)

        except IntegrityError as ie:
            return JsonResponse({'error': f'Error de integridad: {str(ie)}'}, status=500)
        except Exception as e:
            return JsonResponse({'error': f'Error al guardar los datos: {str(e)}'}, status=500)
    else:
        return HttpResponse(status=405)





def generar_reporte_json(request):
    try:
        if request.method == 'GET':
            # Obtener par√°metros de la solicitud
            concepto_nivel_1 = request.GET.get('concepto_nivel_1', 'todos')
            concepto_nivel_2 = request.GET.get('concepto_nivel_2', 'todos')
            fecha_inicio = request.GET.get('fecha_inicio', None)
            fecha_final = request.GET.get('fecha_final', None)

            # Convertir las fechas a objetos datetime
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
                fecha_final_dt = datetime.strptime(fecha_final, "%Y-%m-%d")
            except ValueError:
                return JsonResponse({'error': 'Las fechas proporcionadas no tienen el formato correcto (YYYY-MM-DD).'}, status=400)

            # Filtrar gastos y rendiciones seg√∫n los par√°metros
            gastos = Gasto.objects.filter(fecha_gasto__gte=fecha_inicio_dt, fecha_gasto__lte=fecha_final_dt)
            rendiciones = Rendicion.objects.filter(fecha_operacion__gte=fecha_inicio_dt, fecha_operacion__lte=fecha_final_dt)

            if concepto_nivel_1 != 'todos':
                gastos = gastos.filter(concepto_nivel_1=concepto_nivel_1)
                rendiciones = rendiciones.filter(concepto_nivel_1=concepto_nivel_1)
            if concepto_nivel_2 != 'todos':
                gastos = gastos.filter(concepto_nivel_2=concepto_nivel_2)
                rendiciones = rendiciones.filter(concepto_nivel_2=concepto_nivel_2)

            # Verificar si hay datos para mostrar
            if not gastos.exists() and not rendiciones.exists():
                return JsonResponse({'error': 'No se encontraron registros con los par√°metros seleccionados'}, status=404)

            # Agrupar gastos y rendiciones por concepto_nivel_2
            grouped_data = {}
            total_importe = 0

            for item in list(gastos) + list(rendiciones):
                concepto_2 = item.concepto_nivel_2.concepto_nombre if item.concepto_nivel_2 else "Sin Nivel 2"
                if concepto_2 not in grouped_data:
                    grouped_data[concepto_2] = []

                grouped_data[concepto_2].append({
                    'proveedor': item.nombre_proveedor.razon_social if hasattr(item, 'nombre_proveedor') and item.nombre_proveedor else (
                        item.proveedor.razon_social if hasattr(item, 'proveedor') and item.proveedor else 'Sin proveedor'
                    ),
                    'concepto': item.concepto_nivel_1.concepto_nombre if item.concepto_nivel_1 else 'Sin concepto',
                    'forma_pago': item.tipo_comprobante,
                    'importe': item.importe,
                    'fecha': item.fecha_gasto.strftime("%d/%m/%Y") if hasattr(item, 'fecha_gasto') and item.fecha_gasto else (
                        item.fecha_operacion.strftime("%d/%m/%Y") if item.fecha_operacion else 'Sin Fecha'
                    ),
                    'observacion': item.observacion if hasattr(item, 'observacion') and item.observacion else ''

                })
                total_importe += item.importe if item.importe else 0

            # Generar la fecha actual para el reporte
            fecha_actual = datetime.now().strftime("%d/%m/%Y")

            # Preparar los datos que se enviar√°n
            response_data = {
                'fecha': fecha_actual,
                'gastos_por_grupo': [
                    {'nivel_2': nivel_2, 'gastos': data}
                    for nivel_2, data in grouped_data.items()
                ],
                'total_importe': total_importe
            }

            return JsonResponse(response_data)

    except Exception as e:
        # Si ocurre un error, captura la excepci√≥n y retorna un error 500 con detalles
        return JsonResponse({'error': f'Ocurri√≥ un error al generar el reporte: {str(e)}'}, status=500)



def guardar_proveedor(request):
    if request.method == 'POST':
        ruc_dni = request.POST.get('ruc_dni')
        razon_social = request.POST.get('razon_social')
        nombre_comercial = request.POST.get('nombre_comercial', '').strip() or None
        telefono = request.POST.get('telefono')
        nombre_contacto = request.POST.get('nombre_contacto')

        # Validar campos obligatorios
        if not ruc_dni or not razon_social:
            messages.error(request, 'RUC/DNI y Raz√≥n Social son obligatorios.')
            return redirect('proveedores')

        try:
            Proveedor.objects.create(
                ruc_dni=ruc_dni,
                razon_social=razon_social,
                nombre_comercial=nombre_comercial,
                telefono=telefono,
                nombre_contacto=nombre_contacto,
            )
            messages.success(request, 'Proveedor agregado exitosamente.')
        except ValidationError as e:
            messages.error(request, f'Error al guardar el proveedor: {e}')
        except Exception as e:
            messages.error(request, 'Ocurri√≥ un error inesperado al agregar el proveedor.')

    return redirect('proveedores')


def guardar_cuenta_bancaria(request):
    if request.method == 'POST':
        # Obtener los datos del formulario
        proveedor_id = request.POST.get('proveedor_id')
        nombre_banco = request.POST.get('nombre_banco')
        numero_cuenta = request.POST.get('numero_cuenta')
        tipo_cuenta = request.POST.get('tipo_cuenta')
        cci = request.POST.get('cci', '').strip() or None  # Obtener el CCI del formulario

        try:
            # Obtener el proveedor al que se asociar√° la cuenta bancaria
            proveedor = Proveedor.objects.get(id=proveedor_id)

            # Crear la nueva cuenta bancaria
            cuenta_bancaria = CuentaBancaria(
                proveedor=proveedor,
                nombre_banco=nombre_banco,
                numero_cuenta=numero_cuenta,
                tipo_cuenta=tipo_cuenta,
                cci=cci  # Guardar el CCI
            )
            cuenta_bancaria.save()

            # Mensaje de √©xito
            messages.success(request, f'La cuenta bancaria para {proveedor.razon_social} se ha guardado correctamente.')

        except Proveedor.DoesNotExist:
            messages.error(request, 'El proveedor especificado no existe.')
        except Exception as e:
            messages.error(request, f'Error al guardar la cuenta bancaria: {e}')

        # Redirigir a la p√°gina de proveedores o a la lista que prefieras
        return redirect('proveedores')

    return HttpResponse(status=405)  # M√©todo no permitido si no es un POST



def editar_proveedor(request):
    if request.method == 'POST':
        proveedor_id = request.POST.get('id')
        ruc_dni = request.POST.get('ruc_dni')
        razon_social = request.POST.get('razon_social')
        nombre_comercial = request.POST.get('nombre_comercial', '').strip() or None
        telefono = request.POST.get('telefono')
        nombre_contacto = request.POST.get('nombre_contacto')

        # Obtener el proveedor existente
        proveedor = get_object_or_404(Proveedor, id=proveedor_id)

        # Validar campos obligatorios
        if not ruc_dni or not razon_social:
            messages.error(request, 'RUC/DNI y Raz√≥n Social son obligatorios.')
            return redirect('proveedores')

        try:
            proveedor.ruc_dni = ruc_dni
            proveedor.razon_social = razon_social
            proveedor.nombre_comercial = nombre_comercial
            proveedor.telefono = telefono
            proveedor.nombre_contacto = nombre_contacto
            proveedor.save()
            messages.success(request, 'Proveedor actualizado exitosamente.')
        except ValidationError as e:
            messages.error(request, f'Error al actualizar el proveedor: {e}')
        except Exception as e:
            messages.error(request, 'Ocurri√≥ un error inesperado al actualizar el proveedor.')

    return redirect('proveedores')


def logout_view(request):
    logout(request)
    return redirect('login')

def registrar_rendiciones(request, gasto_id):
    gasto = get_object_or_404(Gasto, id=gasto_id)
    concepto_nivel_1 = Concepto.objects.filter(nivel=1)
    concepto_nivel_2 = Concepto.objects.filter(nivel=2)
    concepto_nivel_3 = Concepto.objects.filter(nivel=3)
    proveedores = Proveedor.objects.all()  # Obtener todos los proveedores

    return render(request, 'gastosRendiciones.html', {
        'gasto': gasto,
        'conceptos_nivel_1': concepto_nivel_1,
        'conceptos_nivel_2': concepto_nivel_2,
        'conceptos_nivel_3': concepto_nivel_3,
        'proveedores': proveedores,  # Pasar proveedores al template
    })





def obtener_saldo_inicial_manual(fecha_inicio, usuario=None):
    dias_busqueda = 5  # M√°ximo de d√≠as para retroceder
    fecha_actual = fecha_inicio - timedelta(days=1)  # Comenzar con el d√≠a anterior al rango

    saldo_final = Decimal(0.00)

    for _ in range(dias_busqueda):
        # Filtrar los ingresos y gastos hasta la fecha actual
        ingresos = Ingreso.objects.filter(fecha_ingreso__lte=fecha_actual)
        gastos = Gasto.objects.filter(fecha_gasto__lte=fecha_actual)

        if usuario:  # Si se proporciona un usuario, filtrar por el creador
            ingresos = ingresos.filter(usuario_creador=usuario)
            gastos = gastos.filter(usuario_creador=usuario)

        # Calcular la suma total de ingresos y gastos
        total_ingresos = ingresos.aggregate(total=Sum('importe'))['total'] or Decimal(0.00)
        total_gastos = gastos.aggregate(total=Sum('importe'))['total'] or Decimal(0.00)

        saldo_final = total_ingresos - total_gastos

        # Si encontramos datos significativos (saldo no nulo), devolvemos el saldo calculado
        if saldo_final != 0:
            return saldo_final

        # Retroceder un d√≠a m√°s
        fecha_actual -= timedelta(days=1)

    # Si no se encontr√≥ saldo en el rango permitido, retornar 0
    return Decimal(0.00)

def actualizar_movimiento(request):
    if request.method == 'POST':
        item_id = request.POST.get('id')
        observacion = request.POST.get('notas')
        tipo = request.POST.get('tipo')  # Se recibe el tipo de movimiento

        try:
            item_id = int(item_id)
        except (ValueError, TypeError):
            messages.error(request, "ID inv√°lido.")
            return redirect('caja_chica')

        if tipo == "Gasto":
            item = get_object_or_404(Gasto, id=item_id)
            # Actualizar la observaci√≥n
            item.observacion = observacion

            # Actualizar conceptos seg√∫n los nombres en el formulario
            concepto1_id = request.POST.get('concepto_nivel_1', '').strip()
            concepto2_id = request.POST.get('concepto_nivel_2', '').strip()
            concepto3_id = request.POST.get('concepto_nivel_3', '').strip()

            if concepto1_id:
                try:
                    item.concepto_nivel_1 = Concepto.objects.get(id=int(concepto1_id))
                except (Concepto.DoesNotExist, ValueError):
                    item.concepto_nivel_1 = None
            else:
                item.concepto_nivel_1 = None

            if concepto2_id:
                try:
                    item.concepto_nivel_2 = Concepto.objects.get(id=int(concepto2_id))
                except (Concepto.DoesNotExist, ValueError):
                    item.concepto_nivel_2 = None
            else:
                item.concepto_nivel_2 = None

            if concepto3_id:
                try:
                    item.concepto_nivel_3 = Concepto.objects.get(id=int(concepto3_id))
                except (Concepto.DoesNotExist, ValueError):
                    item.concepto_nivel_3 = None
            else:
                item.concepto_nivel_3 = None

            item.save()
        else:
            item = get_object_or_404(Ingreso, id=item_id)
            item.observacion = observacion
            item.save()

        messages.success(request, f"Se actualiz√≥ la observaci√≥n del {tipo} con ID {item.id}.")
        return redirect('caja_chica')
    else:
        return JsonResponse({'error': 'M√©todo no permitido'}, status=405)



def descargar_excel(request):
    # Configuraci√≥n inicial y obtenci√≥n de fechas
    hoy = date.today().strftime('%Y-%m-%d')
    fecha_inicio = request.GET.get('fecha_inicio', hoy)
    fecha_fin = request.GET.get('fecha_fin', hoy)
    fecha_inicio = date.fromisoformat(fecha_inicio)
    fecha_fin = date.fromisoformat(fecha_fin)

    # Obtenci√≥n del saldo inicial
    saldo_base = Decimal(0)
    try:
        saldo_inicial = SaldoInicial.objects.get(usuario=request.user)
        saldo_base = saldo_inicial.monto_saldo_inicial
    except SaldoInicial.DoesNotExist:
        saldo_base = Decimal('0.00')

    saldo_inicial = obtener_saldo_inicial_manual(fecha_inicio, usuario=request.user if not request.user.is_staff else None)
    saldo_inicial += saldo_base

    # Obtenci√≥n de datos
    if request.user.is_staff:
        ingresos = Ingreso.objects.filter(
            fecha_ingreso__range=[fecha_inicio, fecha_fin],
            usuario_creador=request.user
        )
        gastos = Gasto.objects.filter(fecha_gasto__range=[fecha_inicio, fecha_fin])
    else:
        ingresos = Ingreso.objects.filter(
            fecha_ingreso__range=[fecha_inicio, fecha_fin],
            usuario_creador=request.user
        )
        gastos = Gasto.objects.filter(
            fecha_gasto__range=[fecha_inicio, fecha_fin],
            usuario_creador=request.user
        )

    # Procesamiento de ingresos
    ingresos_list = []
    for ingreso in ingresos:
        nombre_fondo = ingreso.id_fondo.nombre_fondo if ingreso.id_fondo else 'Sin nombre'
        tipo = 'Extorno' if ingreso.extorno else 'Ingreso'
        ingresos_list.append({
            'tipo': tipo,
            'fecha': ingreso.fecha_ingreso.strftime('%-d/%-m/%Y') if ingreso.fecha_ingreso else '',
            'metodo_pago': ingreso.metodo_pago or '',
            'concepto': nombre_fondo,
            'proveedor': ingreso.id_fondo.nombre_fondo if ingreso.id_fondo else '',
            'banco': ingreso.banco.nombre if ingreso.banco else '',
            'codigo_operacion': ingreso.codigo_operacion or '',
            'fecha_operacion': ingreso.fecha_operacion.strftime('%d/%m/%Y') if ingreso.fecha_operacion else '',
            'notas': ingreso.observacion or '',
            'monto': Decimal(ingreso.importe) if ingreso.importe else Decimal('0.00'),
            'local': ingreso.local.nombre_local if ingreso.local else '-',
            'Comprobante': '-',
            'NumeroComprobante': '-',
        })

    # Procesamiento de gastos
    gastos_list = []
    for gasto in gastos:
        if gasto.concepto_nivel_3:
            concepto = gasto.concepto_nivel_3.concepto_nombre
        elif gasto.concepto_nivel_2:
            concepto = gasto.concepto_nivel_2.concepto_nombre
        elif gasto.concepto_nivel_1:
            concepto = gasto.concepto_nivel_1.concepto_nombre
        else:
            concepto = f"REQ N¬∞{gasto.num_requerimiento} (Id={gasto.id_requerimiento})" if gasto.id_requerimiento else gasto.tipo_comprobante or ''

        gastos_list.append({
            'tipo': 'Gasto',
            'fecha': gasto.fecha_gasto.strftime('%-d/%-m/%Y'),
            'metodo_pago': gasto.tipo_pago,
            'concepto': concepto,
            'proveedor': gasto.nombre_proveedor.razon_social,
            'banco': gasto.banco.nombre if gasto.banco else '',
            'codigo_operacion': gasto.codigo_operacion or '',
            'fecha_operacion': gasto.fecha_operacion.strftime('%d/%m/%Y') if gasto.fecha_operacion else '',
            'notas': gasto.observacion or '',
            'monto': Decimal(gasto.importe),
            'local': gasto.local.nombre_local if gasto.local else '-',
            'Comprobante': gasto.tipo_comprobante or '-',
            'NumeroComprobante': gasto.num_comprobante or '-',
        })

    # Configuraci√≥n de estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_font = Font(bold=True)

    # Crear archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="movimientos_{fecha_inicio}_a_{fecha_fin}.xlsx"'
    wb = Workbook()
    ws = wb.active

    # Escribir encabezados
    ws.append([f"Movimientos de {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"])
    ws.append([])
    ws.append([f'Saldo inicial: {saldo_inicial:,.2f}'])
    ws.append([])

    encabezados = [
        'Tipo', 'Fecha', 'Local', 'M√©todo de Pago', 'Comprobante',
        'Numero de Comprobante', 'Concepto', 'Proveedor', 'Banco',
        'C√≥digo de Operaci√≥n', 'Fecha de Operaci√≥n', 'Monto', 'Notas'
    ]

    for col_num, header in enumerate(encabezados, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    row_num = 6  # Fila inicial para datos

    # Escribir ingresos
    for ingreso in ingresos_list:
        ws.cell(row=row_num, column=1, value=ingreso['tipo'])
        ws.cell(row=row_num, column=2, value=ingreso['fecha'])
        ws.cell(row=row_num, column=3, value=ingreso['local'])
        ws.cell(row=row_num, column=4, value=ingreso['metodo_pago'])
        ws.cell(row=row_num, column=5, value=ingreso['Comprobante'])
        ws.cell(row=row_num, column=6, value=ingreso['NumeroComprobante'])
        ws.cell(row=row_num, column=7, value=ingreso['concepto'])
        ws.cell(row=row_num, column=8, value=ingreso['proveedor'])
        ws.cell(row=row_num, column=9, value=ingreso['banco'])
        ws.cell(row=row_num, column=10, value=ingreso['codigo_operacion'])
        ws.cell(row=row_num, column=11, value=ingreso['fecha_operacion'])
        ws.cell(row=row_num, column=12, value=float(ingreso['monto']))
        ws.cell(row=row_num, column=13, value=ingreso['notas'])
        row_num += 1

    # Total Ingresos
    if ingresos_list:
        total_ingresos = sum(i['monto'] for i in ingresos_list)
        ws.cell(row=row_num, column=1, value='Total Ingresos').font = bold_font
        ws.cell(row=row_num, column=12, value=float(total_ingresos)).fill = green_fill
        ws.cell(row=row_num, column=12).font = bold_font
        row_num += 1

    row_num += 2  # Espacio entre secciones

    # Escribir gastos
    for gasto in gastos_list:
        ws.cell(row=row_num, column=1, value=gasto['tipo'])
        ws.cell(row=row_num, column=2, value=gasto['fecha'])
        ws.cell(row=row_num, column=3, value=gasto['local'])
        ws.cell(row=row_num, column=4, value=gasto['metodo_pago'])
        ws.cell(row=row_num, column=5, value=gasto['Comprobante'])
        ws.cell(row=row_num, column=6, value=gasto['NumeroComprobante'])
        ws.cell(row=row_num, column=7, value=gasto['concepto'])
        ws.cell(row=row_num, column=8, value=gasto['proveedor'])
        ws.cell(row=row_num, column=9, value=gasto['banco'])
        ws.cell(row=row_num, column=10, value=gasto['codigo_operacion'])
        ws.cell(row=row_num, column=11, value=gasto['fecha_operacion'])
        ws.cell(row=row_num, column=12, value=float(gasto['monto']))
        ws.cell(row=row_num, column=13, value=gasto['notas'])
        row_num += 1

    # Total Gastos
    if gastos_list:
        total_gastos = sum(g['monto'] for g in gastos_list)
        ws.cell(row=row_num, column=1, value='Total Gastos').font = bold_font
        ws.cell(row=row_num, column=12, value=float(total_gastos)).fill = red_fill
        ws.cell(row=row_num, column=12).font = bold_font
        row_num += 1

    # Saldo Final
    total_ingresos = sum(i['monto'] for i in ingresos_list) if ingresos_list else Decimal('0.00')
    total_gastos = sum(g['monto'] for g in gastos_list) if gastos_list else Decimal('0.00')
    saldo_final = saldo_inicial + total_ingresos - total_gastos
    ws.append([])
    ws.append([f'Saldo final: {saldo_final:,.2f}'])

    # Ajustar anchos de columnas
    for col in range(1, 14):
        max_length = 0
        column = chr(64 + col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response






def transferir_yape(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            item_id = data.get("item_id")
            monto = data.get("monto")

            # Validar que `item_id` y `monto` existen
            if not item_id or monto is None:
                return JsonResponse({"error": "Faltan datos en la solicitud."}, status=400)

            # Convertir monto a Decimal
            monto = Decimal(monto)

            # Obtener el ingreso
            ingreso = get_object_or_404(Ingreso, id=item_id)

            # Verificar que el monto no sea mayor que el importe disponible
            if monto > ingreso.importe:
                return JsonResponse({"error": "El monto a transferir no puede ser mayor al importe disponible."}, status=400)

            # Actualizar el campo `importe_yape`
            ingreso.importe_yape += monto

            # Calcular la diferencia y actualizar `importe_efectivo`
            ingreso.importe_efectivo = ingreso.importe - ingreso.importe_yape

            ingreso.save()

            return JsonResponse({"message": f"Transferencia exitosa. ID: {item_id}, Monto transferido: {monto}, Nuevo efectivo: {ingreso.importe_efectivo}"})

        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al procesar la solicitud. JSON inv√°lido."}, status=400)
        except Ingreso.DoesNotExist:
            return JsonResponse({"error": "Ingreso no encontrado."}, status=404)
        except ValueError:
            return JsonResponse({"error": "Monto inv√°lido, debe ser un n√∫mero."}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error inesperado: {str(e)}"}, status=500)

    return JsonResponse({"error": "M√©todo no permitido."}, status=405)

from django.views.decorators.http import require_POST

@login_required
@require_POST
def eliminar_item(request, item_id, item_tipo):
    if item_tipo == "Gasto":
        # Buscar el gasto por id y eliminarlo
        item = get_object_or_404(Gasto, id=item_id)
        item.delete()
    else:
        # Si no es Gasto se asume que es Ingreso, se busca y elimina
        item = get_object_or_404(Ingreso, id=item_id)
        item.delete()

    messages.success(request, f"Eliminado correctamente: ID {item_id}, Tipo {item_tipo}")
    return redirect("caja_chica")


@login_required
def caja_chica(request):
    hoy = date.today().strftime('%Y-%m-%d')  # Formato para los campos de tipo date
    usuario = request.user
    hoy1 = date.today()

    saldo_inicial = usuario.saldo_inicial  # Gracias a related_name="saldo_inicial"
    if saldo_inicial.fecha_cierre and saldo_inicial.fecha_cierre < hoy1:
        saldo_inicial.caja_cerrada = False  # Cerrar caja si la fecha de cierre es antigua
        saldo_inicial.save()

    # Obtener las fechas de los par√°metros GET, si no est√°n presentes usar la fecha de hoy
    fecha_inicio = request.GET.get('fecha_inicio', hoy)
    fecha_fin = request.GET.get('fecha_fin', hoy)

    # Convertir las fechas a tipo date
    fecha_inicio = date.fromisoformat(fecha_inicio)
    fecha_fin = date.fromisoformat(fecha_fin)

    # Calcular el saldo base seg√∫n el username del usuario autenticado
    saldo_base = Decimal(0)
    try:
        saldo_inicial = SaldoInicial.objects.get(usuario=request.user)
        saldo_base = saldo_inicial.monto_saldo_inicial
        saldo_yape = saldo_inicial.monto_saldo_inicial_yape
    except SaldoInicial.DoesNotExist:
        saldo_base = Decimal('0.00')
        saldo_yape = Decimal('0.00')

    # Calcular el saldo inicial usando la funci√≥n personalizada
    saldo_inicial = obtener_saldo_inicial_manual(fecha_inicio, usuario=request.user if not request.user.is_staff else None)
    # Sumar el saldo base al saldo inicial
    saldo_inicial += saldo_base
    saldo_inicial += saldo_yape
    saldo_efectivo = 0
    saldo_banco = 0
    if not request.user.is_staff:
        # Filtrar los ingresos y gastos del usuario autenticado sin l√≠mite de fechas
        ingresos = Ingreso.objects.filter(usuario_creador=request.user).exclude(metodo_pago="Sin especificar")
        gastos = Gasto.objects.filter(usuario_creador=request.user)

        # Separar ingresos en dos listas seg√∫n importe_yape
        ingresos_efectivo = [ingreso.importe if ingreso.importe_yape == 0 else ingreso.importe_efectivo for ingreso in ingresos if ingreso.metodo_pago == "efectivo"]
        ingresos_banco = [ingreso.importe for ingreso in ingresos if ingreso.metodo_pago != "efectivo"]
        ingresos_yape = [ingreso.importe_yape for ingreso in ingresos if ingreso.importe_yape != 0]

        # Separar gastos seg√∫n el tipo de pago
        gastos_efectivo = [gasto.importe for gasto in gastos if gasto.tipo_pago == "efectivo"]
        gastos_banco = [gasto.importe for gasto in gastos if gasto.tipo_pago != "efectivo"]

        # Calcular saldo efectivo
        saldo_efectivo = saldo_base + sum(ingresos_efectivo) - sum(gastos_efectivo)

        # Calcular saldo banco
        saldo_banco = saldo_yape + sum(ingresos_banco) + sum(ingresos_yape) - sum(gastos_banco)



    # Verificar si el usuario es staff
    if request.user.is_staff:
        # Staff puede ver todos los ingresos y gastos
        ingresos = Ingreso.objects.filter(
            fecha_ingreso__range=[fecha_inicio, fecha_fin],
            usuario_creador=request.user  # Filtra solo los ingresos creados por el usuario staff
        )
        gastos = Gasto.objects.filter(fecha_gasto__range=[fecha_inicio, fecha_fin])
    else:
        # No staff solo puede ver los ingresos y gastos que cre√≥
        ingresos = Ingreso.objects.filter(
            fecha_ingreso__range=[fecha_inicio, fecha_fin],
            usuario_creador=request.user
        )
        gastos = Gasto.objects.filter(
            fecha_gasto__range=[fecha_inicio, fecha_fin],
            usuario_creador=request.user
        )

    # Preparar los movimientos para la tabla
    movimientos = []

    # Procesar ingresos
    for ingreso in ingresos:
        nombre_fondo = ingreso.id_fondo.nombre_fondo if ingreso.id_fondo else 'Sin nombre'
        tipo = 'Extorno' if ingreso.extorno else 'Ingreso'
        movimientos.append({
            'id': ingreso.id,
            'tipo': tipo,
            'fecha': ingreso.fecha_ingreso.strftime('%-d/%-m/%Y'),
            'metodo_pago': ingreso.metodo_pago,
            'concepto': ingreso.local.nombre_local if ingreso.local else "sin local asignado",
            'notas': ingreso.observacion or '',
            'monto': Decimal(ingreso.importe),
            'moneda': ingreso.moneda,
            'proveedor': ingreso.id_fondo.nombre_fondo if ingreso.id_fondo else "Sin fondo especifico",
            'usuario_creador': ingreso.usuario_creador.username if ingreso.usuario_creador else "Desconocido",
            'transferir_yape':ingreso.importe_yape,
            'eliminar': "Si" if ingreso.fecha_ingreso == hoy1 else "No"

        })

    # Procesar gastos
    for gasto in gastos:
        if gasto.concepto_nivel_3:
            concepto = gasto.concepto_nivel_3.concepto_nombre
        elif gasto.concepto_nivel_2:
            concepto = gasto.concepto_nivel_2.concepto_nombre
        elif gasto.concepto_nivel_1:
            concepto = gasto.concepto_nivel_1.concepto_nombre
        else:
            if gasto.id_requerimiento and gasto.num_requerimiento:
                concepto = f"REQ N¬∞{gasto.num_requerimiento} (Id={gasto.id_requerimiento})"
            else:
                concepto = gasto.tipo_comprobante or ''
        total_rendido = sum(rendicion.importe for rendicion in gasto.rendiciones_gasto.all())

        movimientos.append({
            'id': gasto.id,
            'tipo': 'Gasto',
            'fecha': gasto.fecha_gasto.strftime('%-d/%-m/%Y'),
            'metodo_pago': gasto.tipo_pago,
            'concepto': concepto,
            'notas': gasto.observacion or '',
            'monto': Decimal(gasto.importe),
            'rendido': gasto.rendido,
            'moneda': gasto.moneda,
            'proveedor': gasto.nombre_proveedor.razon_social if gasto.nombre_proveedor and gasto.nombre_proveedor.razon_social else "No hay raz√≥n social",
            'rendiciones': list(gasto.rendiciones_gasto.all()),
            'total_rendido': total_rendido,  # ‚úÖ Agregar el total de rendiciones
            'usuario_creador': gasto.usuario_creador.username,
            'concepto_nivel_1_id': gasto.concepto_nivel_1.id if gasto.concepto_nivel_1 else "",
            'concepto_nivel_2_id': gasto.concepto_nivel_2.id if gasto.concepto_nivel_2 else "",
            'concepto_nivel_3_id': gasto.concepto_nivel_3.id if gasto.concepto_nivel_3 else "",
            'tipo_item':gasto.tipo_comprobante,
            'eliminar': "Si" if gasto.fecha_gasto == hoy1 else "No"

        })

    # C√°lculos para los totales
    total_ingresos_soles = sum(m['monto'] for m in movimientos if m['tipo'] != 'Gasto' and m['moneda'] == 'Soles')
    total_egresos_soles = sum(m['monto'] for m in movimientos if m['tipo'] == 'Gasto' and m['moneda'] == 'Soles')
    saldo_final = saldo_inicial + total_ingresos_soles - total_egresos_soles

    # Generar el t√≠tulo para las fechas seleccionadas
    rango_fechas = f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    nivel_1_conceptos = Concepto.objects.filter(nivel=1)
    nivel_2_conceptos = Concepto.objects.filter(nivel=2)
    nivel_3_conceptos = Concepto.objects.filter(nivel=3)
    # Contexto para la plantilla
    context = {
        'movimientos': movimientos,
        'fecha_hoy': hoy,  # Pasar la fecha actual
        'saldo_inicial': saldo_inicial,
        'total_ingresos_soles': total_ingresos_soles,
        'total_egresos_soles': total_egresos_soles,
        'saldo_final': saldo_final,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'rango_fechas': rango_fechas,
        'saldo_efectivo': saldo_efectivo,
        'saldo_banco': saldo_banco,
        'nivel_1_conceptos': nivel_1_conceptos,
        'nivel_2_conceptos': nivel_2_conceptos,
        'nivel_3_conceptos': nivel_3_conceptos,
    }

    return render(request, 'caja.html', context)









def verificar_gastos(request, proveedor_id):
    """Verifica si un proveedor tiene gastos antes de eliminarlo"""
    tiene_gastos = Gasto.objects.filter(nombre_proveedor_id=proveedor_id).exists()
    return JsonResponse({'tiene_gastos': tiene_gastos})

def cuentas_bancarias(request, proveedor_id):
    proveedor = Proveedor.objects.get(id=proveedor_id)
    cuentas = CuentaBancaria.objects.filter(proveedor_id=proveedor_id)
    cuentas_data = [{
        'nombre_banco': cuenta.nombre_banco,
        'numero_cuenta': cuenta.numero_cuenta,
        'tipo_cuenta': cuenta.get_tipo_cuenta_display(),
        'cci': cuenta.cci,
    } for cuenta in cuentas]
    return JsonResponse({'proveedor': proveedor.razon_social, 'cuentas': cuentas_data})


def prueba(request):
    return render(request, 'prueba.html')
@login_required
def proveedores(request):
    proveedores = Proveedor.objects.all()
    return render(request, 'proveedores.html', {'proveedores': proveedores})
@login_required
def eliminar_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    proveedor.delete()
    messages.success(request, 'Proveedor eliminado exitosamente.')
    return redirect('proveedores')


def validar_campos(metodo_pago, data):
    if metodo_pago != "efectivo":
        if not data.get("codigo_operacion"):
            return "El c√≥digo de operaci√≥n es obligatorio."
        if not data.get("fecha_operacion"):
            return "La fecha de operaci√≥n es obligatoria."
        if not data.get("banco_operacion"):
            return "El banco es obligatorio."
    return None

@login_required
@csrf_exempt
def editar_ingreso(request, id):
    ingreso = get_object_or_404(Ingreso, id=id)

    if request.method == "POST":
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"message": "Datos JSON inv√°lidos."}, status=400)

        metodo_pago = data.get("metodo_pago")
        error_msg = validar_campos(metodo_pago, data)
        if error_msg:
            return JsonResponse({"message": error_msg}, status=400)

        ingreso.id_fondo = get_object_or_404(Fondo, id=data['id_fondo'])
        ingreso.metodo_pago = metodo_pago
        ingreso.local = get_object_or_404(Local, id=data['local'])

        # Solo asignar codigo_operacion, fecha_operacion y banco si el metodo_pago no es "efectivo"
        if metodo_pago != "efectivo":
            ingreso.codigo_operacion = data.get('codigo_operacion')
            ingreso.fecha_operacion = data.get('fecha_operacion')
            ingreso.banco = get_object_or_404(Banco, id=data['banco_operacion']) if data.get('banco_operacion') else None

        ingreso.observacion = data['observacion']
        ingreso.save()

        return JsonResponse({"message": "El ingreso ha sido actualizado exitosamente."}, status=200)

    return JsonResponse({"message": "M√©todo no permitido"}, status=405)


@login_required
def edit_item(request, id, tipo):
    nivel_1_conceptos = Concepto.objects.filter(nivel=1)
    nivel_2_conceptos = Concepto.objects.filter(nivel=2)
    nivel_3_conceptos = Concepto.objects.filter(nivel=3)
    fondos = Fondo.objects.all()
    locales = Local.objects.all()
    bancos = Banco.objects.all()
    if tipo == 'Extorno':
        ingreso = get_object_or_404(Ingreso, id=id)
                # Convertir importe a cadena con punto decimal
        if ingreso.importe is not None:
            ingreso.importe = f"{float(ingreso.importe):.2f}"
        return render(request, 'edit_ingreso.html', {'ingreso': ingreso,
                    'fondos': fondos,
                    'bancos':bancos,
        'locales': locales,
        })
    elif tipo == 'Gasto':
        gasto = get_object_or_404(Gasto, id=id)

        # Convertir importe a cadena con punto decimal
        if gasto.importe is not None:
            gasto.importe = f"{float(gasto.importe):.2f}"

        return render(request, 'edit_gasto.html', {'gasto': gasto,
                    'nivel_1_conceptos': nivel_1_conceptos,
        'nivel_2_conceptos': nivel_2_conceptos,
        'nivel_3_conceptos': nivel_3_conceptos,
        'fondos': fondos,
        'locales': locales,
        })
def buscar_proveedores(request):
    search_term = request.GET.get('search', '')
    proveedores = Proveedor.objects.filter(
        razon_social__icontains=search_term
    ) | Proveedor.objects.filter(
        nombre_comercial__icontains=search_term
    )
    data = [
        {
            "id": proveedor.id,
            "razon_social": proveedor.razon_social,
            "nombre_comercial": proveedor.nombre_comercial,
        }
        for proveedor in proveedores
    ]
    return JsonResponse(data, safe=False)

def registrar_gasto(request, id):
    gasto = get_object_or_404(Gasto, id=id)

    if request.method == "POST":
        try:
            # Actualizar los campos con los valores enviados por el formulario
            gasto.tipo_pago = request.POST.get('metodo_pago')
            gasto.concepto_nivel_1_id = request.POST.get('concepto_nivel_1_results') or None
            gasto.concepto_nivel_2_id = request.POST.get('concepto_nivel_2_results') or None
            gasto.concepto_nivel_3_id = request.POST.get('concepto_nivel_3_results') or None
            gasto.nombre_proveedor = request.POST.get('nombre_proveedor', '').strip()
            local_id = request.POST.get('local')

            gasto.importe = request.POST.get('importe') or 0
            gasto.observacion = request.POST.get('observacion', '')
            if gasto.tipo_pago in ['transferencia', 'yape']:
                gasto.codigo_operacion = request.POST.get('codigo_operacion', '').strip()
                gasto.fecha_operacion = request.POST.get('fecha_operacion') or None
            # Obtener y asignar el local seleccionado
            if local_id:
                gasto.local_id = local_id
            gasto.usuario_creador = request.user

            gasto.save()

            # Mensaje de √©xito
            messages.success(request, "Gasto registrado correctamente.")
            return redirect('caja_chica')  # Redirigir a la URL caja_chica

        except Exception as e:
            messages.error(request, f"Ocurri√≥ un error al registrar el gasto: {str(e)}")
            return redirect('gasto_edit', id=id)  # Redirige al formulario original en caso de error

    return redirect('gasto_edit', id=id)

def gasto_edit(request, id):
    gasto = get_object_or_404(Gasto, id=id)

    if request.method == "POST":
        try:
            # Actualizar los campos con los valores enviados por el formulario
            gasto.tipo_pago = request.POST.get('metodo_pago')
            gasto.concepto_nivel_1_id = request.POST.get('concepto_nivel_1_results') or None
            gasto.concepto_nivel_2_id = request.POST.get('concepto_nivel_2_results') or None
            gasto.concepto_nivel_3_id = request.POST.get('concepto_nivel_3_results') or None
            gasto.importe = request.POST.get('importe') or 0
            gasto.observacion = request.POST.get('observacion', '')
            gasto.fecha_operacion = request.POST.get('fecha_operacion') or None
            gasto.codigo_operacion = request.POST.get('codigo_operacion', '')

            # Guardar los cambios en la base de datos
            gasto.save()

            # Mensaje de √©xito
            messages.success(request, "Gasto registrado correctamente.")
            return redirect('caja_chica')  # Redirigir a la URL caja_chica

        except Exception as e:
            # Mensaje de error
            messages.error(request, f"Ocurri√≥ un error al registrar el gasto: {str(e)}")
            return redirect('gasto_edit', id=id)  # Redirige al mismo formulario en caso de error

    # Renderizar el formulario con los datos actuales del gasto
    return render(request, 'edit_gasto.html', {
        'gasto': gasto,
        'nivel_1_conceptos': Concepto.objects.filter(nivel=1),  # Asumiendo una estructura de niveles
        'nivel_2_conceptos': Concepto.objects.filter(nivel=2),
        'nivel_3_conceptos': Concepto.objects.filter(nivel=3),
        'locales': Local.objects.all(),
    })


from django.urls import reverse

def guardar_oficial(request):
    if request.method == "POST":
        try:
            # Parsear el cuerpo de la solicitud como JSON
            body = json.loads(request.body)
            rendiciones = body.get('rendiciones', [])
            gasto_id = body.get('gasto_id')  # ID del gasto recibido

            # Obtener el gasto asociado
            gasto = Gasto.objects.get(id=gasto_id)

            # Procesar los datos recibidos
            data_guardada = []
            suma_importes_rendiciones = Decimal(0)  # Acumulador para la suma de los importes de las rendiciones
            mensaje_creacion = None  # Variable para guardar el mensaje de creaci√≥n

            for rendicion in rendiciones:
                fecha_operacion = rendicion.get('fecha_operacion')
                proveedor = rendicion.get('apellidos_nombres_proveedor')
                numero_requerimiento = rendicion.get('numero_requerimiento')
                descripcion=rendicion.get('descripcion')
                importe = Decimal(rendicion.get('importe', 0))
                concepto_nivel_1_id = rendicion.get('concepto_nivel_1')
                concepto_nivel_2_id = rendicion.get('concepto_nivel_2')
                concepto_nivel_3_id = rendicion.get('concepto_nivel_3')
                tipo_comprobante = rendicion.get('tipo_comprobante')  # Captura del tipo de comprobante

                proveedor = Proveedor.objects.get(id=proveedor)  # Buscar proveedor por ID

                # Crear y guardar objeto Rendicion
                nueva_rendicion = Rendicion(
                    fecha_operacion=fecha_operacion,
                    proveedor=proveedor ,
                    numero_requerimiento=numero_requerimiento,
                    importe=importe,
                    descripcion=descripcion,
                    concepto_nivel_1_id=concepto_nivel_1_id,
                    concepto_nivel_2_id=concepto_nivel_2_id,
                    concepto_nivel_3_id=concepto_nivel_3_id,
                    tipo_comprobante=tipo_comprobante,  # Guardar el tipo de comprobante
                    gasto_id=gasto_id  # Asociar el gasto
                )
                nueva_rendicion.usuario_creador = request.user
                nueva_rendicion.save()

                suma_importes_rendiciones += importe

                data_guardada.append({
                    "id": nueva_rendicion.id,
                    "fecha_operacion": nueva_rendicion.fecha_operacion,
                    "descripcion":nueva_rendicion.descripcion,
                    "proveedor": nueva_rendicion.proveedor,
                    "numero_requerimiento": nueva_rendicion.numero_requerimiento,
                    "importe": nueva_rendicion.importe,
                    "tipo_comprobante": nueva_rendicion.tipo_comprobante,  # Incluir en la respuesta

                })

            # Comparar la suma de los importes de las rendiciones con el importe del gasto
            if suma_importes_rendiciones > gasto.importe:
                # Crear un nuevo gasto por la diferencia
                diferencia = suma_importes_rendiciones - gasto.importe
                nuevo_gasto = Gasto(
                    importe=diferencia,
                    fecha_registro=date.today(),
                    fecha_gasto=date.today(),
                    moneda=gasto.moneda,  # Moneda por defecto del gasto asociado
                    tipo_comprobante=gasto.tipo_comprobante,
                    campo_area=gasto.campo_area,
                    num_requerimiento=gasto.num_requerimiento,
                    id_requerimiento=gasto.id_requerimiento,
                    nombre_proveedor=gasto.nombre_proveedor,
                    local=gasto.local,
                    tipo_pago="efectivo",
                    gasto_origen=gasto,  # Vincular con el gasto original


                )
                nuevo_gasto.usuario_creador = request.user

                nuevo_gasto.save()
                mensaje_creacion = f"Se cre√≥ un nuevo gasto con un importe de {diferencia} y moneda {gasto.moneda}."

            elif suma_importes_rendiciones < gasto.importe:
                # Crear un nuevo ingreso por la diferencia
                diferencia = gasto.importe - suma_importes_rendiciones
                nuevo_ingreso = Ingreso(
                    fecha_registro=date.today(),
                    fecha_ingreso=date.today(),
                    importe=diferencia,
                    metodo_pago="Sin especificar",
                    moneda=gasto.moneda,  # Moneda por defecto del gasto asociado
                    extorno=True,
                    observacion="gasto extra generado",
                    gasto_origen=gasto  # Vincular con el gasto original

                )
                nuevo_ingreso.usuario_creador=request.user
                nuevo_ingreso.save()
                mensaje_creacion = f"Se cre√≥ un nuevo ingreso con un importe de {diferencia} y moneda {gasto.moneda}."

            # Actualizar el campo 'rendido' del gasto asociado
            gasto.rendido = True
            gasto.fecha_rendido = date.today()
            gasto.save()

            # Responder con JSON
            response_data = {
                'status': 'success',
                'message': 'Rendiciones guardadas exitosamente.',
                'detalle': mensaje_creacion,
                'redirect_url': reverse('rendicion')
            }
            return JsonResponse(response_data)

        except Exception as e:
            response_data = {
                'status': 'error',
                'message': f"Error: {str(e)}"
            }
            return JsonResponse(response_data)
    else:
        return redirect('rendicion')




# Create your views here.
def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            if user.username == "naye123":
                return redirect('ver_personal')
            return redirect('dashboard')  # Redirige a la vista del dashboard
        else:
            messages.error(request, "Nombre de usuario o contrase√±a incorrectos.")
            return redirect('login')

    return render(request, 'login.html')


from django.db.models.functions import TruncMonth


from django.http import JsonResponse
from django.db.models import Value
from django.db.models.functions import Coalesce

def calcular_gastos_por_metodo_pago(usuario):
    """
    Calcula la suma de los gastos agrupados por m√©todo de pago.
    Retorna un diccionario con etiquetas (m√©todos de pago) y valores (totales).
    Si el usuario es staff, se mostrar√°n todos los gastos.
    """
    # Si el usuario es staff, no filtramos por usuario_creador
    if usuario.is_staff:
        gastos_por_metodo = (
            Gasto.objects
            .values('tipo_pago')
            .annotate(total_gasto=Sum('importe'))
            .order_by('-total_gasto')
        )
    else:
        # Si no es staff, filtramos por usuario_creador
        gastos_por_metodo = (
            Gasto.objects
            .filter(usuario_creador=usuario)
            .values('tipo_pago')
            .annotate(total_gasto=Sum('importe'))
            .order_by('-total_gasto')
        )

    # Preparar datos para el gr√°fico
    etiquetas = [gasto['tipo_pago'] or 'Sin especificar' for gasto in gastos_por_metodo]
    valores = [float(gasto['total_gasto'] or 0) for gasto in gastos_por_metodo]

    return {
        'etiquetas': etiquetas,
        'valores': valores,
    }

@login_required
def dashboard_view(request):
    if request.user.is_authenticated:
        # Fecha actual
        today = date.today()
        current_year = today.year
        current_month = today.month
        usuario = request.user
        hoy1 = date.today()

        saldo_inicial = usuario.saldo_inicial  # Gracias a related_name="saldo_inicial"
        if saldo_inicial.fecha_cierre and saldo_inicial.fecha_cierre < hoy1:
            saldo_inicial.caja_cerrada = False  # Cerrar caja si la fecha de cierre es antigua
            saldo_inicial.save()
        # Si el usuario es staff, no filtramos los gastos por usuario_creador
        if request.user.is_staff:
            # Si es staff, no filtramos los gastos por usuario logueado
            gastos_filter = Gasto.objects
        else:
            # Si no es staff, filtramos los gastos por usuario_creador
            gastos_filter = Gasto.objects.filter(usuario_creador=request.user)

        # Los ingresos siempre se filtran por el usuario logueado, sin importar si es staff
        ingresos_filter = Ingreso.objects.filter(usuario_creador=request.user)

        # Obtener los 5 conceptos con m√°s gasto (filtrar solo por los gastos, y si es staff ver todos)
        top_conceptos = (
            gastos_filter
            .annotate(concepto_nombre=Coalesce(
                'concepto_nivel_3__concepto_nombre',
                'concepto_nivel_2__concepto_nombre',
                'concepto_nivel_1__concepto_nombre',
                Value('Sin concepto'),
            ))
            .values('concepto_nombre')
            .annotate(total_gasto=Sum('importe'))
            .order_by('-total_gasto')[:5]
        )

        # Etiquetas y valores para el gr√°fico
        conceptos_labels = [entry['concepto_nombre'] for entry in top_conceptos]
        conceptos_data = [float(entry['total_gasto']) for entry in top_conceptos]

        # Obtener ingresos por mes para el gr√°fico de barras (solo mes actual) filtrado por usuario
        ingresos_mes_actual = ingresos_filter.filter(
            fecha_ingreso__year=current_year,
            fecha_ingreso__month=current_month
        )
        ingresos_mes_actual_total = ingresos_mes_actual.aggregate(total_ingresos=Sum('importe'))['total_ingresos'] or Decimal('0.00')

        # Obtener gastos por mes para el gr√°fico de barras (solo mes actual) filtrado por usuario o todos si es staff
        gastos_mes_actual = gastos_filter.filter(
            fecha_gasto__year=current_year,
            fecha_gasto__month=current_month
        )
        gastos_mes_actual_total = gastos_mes_actual.aggregate(total_gastos=Sum('importe'))['total_gastos'] or Decimal('0.00')

        # Obtener ingresos por mes para el gr√°fico de l√≠neas (todo el a√±o actual) filtrado por usuario
        ingresos_anuales = ingresos_filter.filter(
            fecha_ingreso__year=current_year
        ).annotate(month=TruncMonth('fecha_ingreso')).values('month').annotate(total_ingresos=Sum('importe')).order_by('month')

        # Obtener gastos por mes para el gr√°fico de l√≠neas (todo el a√±o actual) filtrado por usuario o todos si es staff
        gastos_anuales = gastos_filter.filter(
            fecha_gasto__year=current_year
        ).annotate(month=TruncMonth('fecha_gasto')).values('month').annotate(total_gastos=Sum('importe')).order_by('month')

        # Formatear los resultados para que sean m√°s f√°ciles de usar en los gr√°ficos
        ingresos_data = {entry['month'].strftime('%B %Y'): float(entry['total_ingresos'] or 0) for entry in ingresos_anuales}
        gastos_data = {entry['month'].strftime('%B %Y'): float(entry['total_gastos'] or 0) for entry in gastos_anuales}

        # Generar las etiquetas y los datos de los gr√°ficos de l√≠neas (para todo el a√±o)
        labels_line = list(set(ingresos_data.keys()).union(gastos_data.keys()))
        ingresos_por_mes_line = [ingresos_data.get(label, 0) for label in labels_line]
        gastos_por_mes_line = [gastos_data.get(label, 0) for label in labels_line]

        # Para el gr√°fico de barras (solo mes actual)
        labels_bar = [today.strftime('%B %Y')]  # Solo el mes actual
        ingresos_por_mes_bar = [float(ingresos_mes_actual_total)]
        gastos_por_mes_bar = [float(gastos_mes_actual_total)]

        # Llamar a la funci√≥n para calcular los gastos por m√©todo de pago (tambi√©n filtrados por usuario)
        datos_gastos_metodo_pago = calcular_gastos_por_metodo_pago(request.user)

        # Enviar los datos a la plantilla
        data = {
            'labels_bar': labels_bar,
            'ingresos_por_mes_bar': ingresos_por_mes_bar,
            'gastos_por_mes_bar': gastos_por_mes_bar,
            'labels_line': labels_line,
            'ingresos_por_mes_line': ingresos_por_mes_line,
            'gastos_por_mes_line': gastos_por_mes_line,
            'conceptos_labels': conceptos_labels,
            'conceptos_data': conceptos_data,
            'datos_gastos_metodo_pago': datos_gastos_metodo_pago,
        }

        return render(request, 'dashboard.html', data)
    else:
        return redirect('login')

def reporte_concepto_proveedor_pdf(request):
    if request.method == "POST":
        # Recoger datos del formulario (inputs de tipo month: formato "YYYY-MM")
        proveedor_id = request.POST.get("proveedor")
        concepto_nivel_1 = request.POST.get("concepto_nivel_1")  # Puede ser 'todos' o un ID
        fecha_inicio = request.POST.get("fecha_inicio")
        fecha_final = request.POST.get("fecha_final")

        # Validar que se reciban todos los datos necesarios
        if not all([proveedor_id, concepto_nivel_1, fecha_inicio, fecha_final]):
            messages.error(request, "Faltan datos para generar el reporte.")
            return redirect("reportes")

        try:
            # Convertir las cadenas "YYYY-MM" en fechas: primer d√≠a del mes y √∫ltimo d√≠a del mes
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m").date().replace(day=1)
            import calendar
            year_final, month_final = map(int, fecha_final.split('-'))
            last_day = calendar.monthrange(year_final, month_final)[1]
            fecha_final_dt = datetime.strptime(fecha_final, "%Y-%m").date().replace(day=last_day)
        except Exception:
            messages.error(request, "Formato de fecha incorrecto.")
            return redirect("reportes")

        try:
            proveedor = Proveedor.objects.get(pk=proveedor_id)
        except Proveedor.DoesNotExist:
            messages.error(request, "Proveedor no encontrado.")
            return redirect("reportes")

        # Obtener registros de Gastos para el proveedor en el rango de fechas
        gastos_qs = Gasto.objects.filter(
            nombre_proveedor=proveedor,
            fecha_gasto__range=[fecha_inicio_dt, fecha_final_dt]
        )
        if concepto_nivel_1 != "todos":
            gastos_qs = gastos_qs.filter(concepto_nivel_1__id=concepto_nivel_1)

        # Obtener registros de Rendiciones para el proveedor en el rango de fechas
        rendiciones_qs = Rendicion.objects.filter(
            proveedor=proveedor,
            fecha_operacion__range=[fecha_inicio_dt, fecha_final_dt]
        )
        if concepto_nivel_1 != "todos":
            rendiciones_qs = rendiciones_qs.filter(concepto_nivel_1__id=concepto_nivel_1)

        # Combinar ambos querysets
        registros = list(gastos_qs) + list(rendiciones_qs)

        # Ordenar registros por fecha (usando fecha_gasto para Gastos y fecha_operacion para Rendiciones)
        def get_registro_fecha(reg):
            if hasattr(reg, "fecha_gasto") and reg.fecha_gasto:
                return reg.fecha_gasto
            elif hasattr(reg, "fecha_operacion") and reg.fecha_operacion:
                return reg.fecha_operacion
            return datetime.today().date()
        registros.sort(key=get_registro_fecha)

        # Determinar el texto del concepto
        if concepto_nivel_1 == "todos":
            concepto_text = "Todos los Conceptos"
        else:
            try:
                concepto_obj = Concepto.objects.get(pk=concepto_nivel_1)
                concepto_text = concepto_obj.concepto_nombre
            except Concepto.DoesNotExist:
                concepto_text = ""

        # Crear el PDF utilizando ReportLab
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Definir un estilo para el Paragraph del √∫ltimo concepto que permita el ajuste de l√≠nea
        concepto_style = ParagraphStyle(
            'ConceptoWrap',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=10,
            leading=12,
            alignment=1  # Centrado
        )

        # T√≠tulo del reporte
        title_text = f"Gasto por concepto '{concepto_text}' del proveedor {proveedor.razon_social}"
        elements.append(Paragraph(title_text, styles['Title']))
        elements.append(Spacer(1, 12))

        # Encabezados de la tabla: Fecha, Monto, √öltimo Concepto, Observaci√≥n/Nota
        data = [["Fecha", "Monto", "Concepto", "Observaciones"]]
        total_importe = 0

        for reg in registros:
            # Obtener la fecha en formato dd/mm/YYYY
            if hasattr(reg, "fecha_gasto") and reg.fecha_gasto:
                fecha_str = reg.fecha_gasto.strftime("%d/%m/%Y")
            elif hasattr(reg, "fecha_operacion") and reg.fecha_operacion:
                fecha_str = reg.fecha_operacion.strftime("%d/%m/%Y")
            else:
                fecha_str = "-"

            # Obtener el monto
            monto_val = float(reg.importe) if reg.importe else 0
            total_importe += monto_val

            # Determinar el √∫ltimo concepto asociado:
            # Se busca primero en nivel 3; si no existe, en nivel 2; de lo contrario, en nivel 1.
            if hasattr(reg, "concepto_nivel_3") and reg.concepto_nivel_3:
                ultimo_concepto = reg.concepto_nivel_3.concepto_nombre
            elif hasattr(reg, "concepto_nivel_2") and reg.concepto_nivel_2:
                ultimo_concepto = reg.concepto_nivel_2.concepto_nombre
            elif hasattr(reg, "concepto_nivel_1") and reg.concepto_nivel_1:
                ultimo_concepto = reg.concepto_nivel_1.concepto_nombre
            else:
                ultimo_concepto = "-"

            # Crear un Paragraph para que el texto se envuelva correctamente
            ultimo_concepto_para = Paragraph(ultimo_concepto, concepto_style)

            # Obtener la observaci√≥n/nota (se usa 'observacion' en Gasto y 'nota' en Rendicion, si existen)
            observacion = "-"
            if hasattr(reg, "observacion") and reg.observacion:
                observacion = reg.observacion
            elif hasattr(reg, "nota") and reg.nota:
                observacion = reg.nota

            data.append([fecha_str, f"{monto_val:,.2f}", ultimo_concepto_para, observacion])

        # Agregar fila de total
        data.append(["Total", f"{total_importe:,.2f}", "", ""])

        # Crear la tabla con estilos y colores bonitos
        table = Table(data, colWidths=[80, 80, 150, 220])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#ADD8E6')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#D1D5DB')),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold')
        ]))
        elements.append(table)

        # Construir el PDF
        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"reporte_concepto_{proveedor.razon_social}_{fecha_inicio}_{fecha_final}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    else:
        messages.error(request, "M√©todo no permitido")
        return redirect("reportes")
def generar_pdf_personal(request):
    sede_nombre = request.GET.get('sede', 'Todas las Sedes')
    planilla_filtro = request.GET.get('planilla', 'Todos')

    # Filtrar personal seg√∫n los par√°metros
    personal = Personal.objects.all()

    if sede_nombre and sede_nombre != "Todas las Sedes":
        personal = personal.filter(local__nombre_local=sede_nombre)

    if planilla_filtro == "Con planilla":
        personal = personal.exclude(regimen_salud="ninguno").exclude(regimen_pensionario="ninguno")
    elif planilla_filtro == "Sin planilla":
        personal = personal.filter(regimen_salud="ninguno", regimen_pensionario="ninguno")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_personal.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), leftMargin=30, rightMargin=30, topMargin=50, bottomMargin=50)
    elements = []

    # Estilos de texto
    styles = getSampleStyleSheet()
    style_table = ParagraphStyle(
        name="TableStyle",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,  # Espaciado entre l√≠neas dentro de la celda
        alignment=1,  # Centrado
    )

    # T√≠tulo con l√≠neas ajustables
    titulo_texto = f"Reporte de Personal<br/>Sede: {sede_nombre} | Planilla: {planilla_filtro}"
    titulo = Paragraph(titulo_texto, styles["Title"])
    elements.append(titulo)
    elements.append(Spacer(1, 0.3 * inch))  # Espacio antes de la tabla

    # Datos de la tabla con ajuste autom√°tico
    data = [["N¬∞ DNI", "Apellidos y Nombres", "Correo Personal", "Celular", "Periodo Inicio", "Periodo Fin", "Sede", "Planilla"]]

    for idx, persona in enumerate(personal, start=1):
        sede = persona.local.nombre_local if persona.local else "Sin Especificar"
        es_planilla = "S√≠" if persona.regimen_salud != "ninguno" and persona.regimen_pensionario != "ninguno" else "No"

        data.append([
            persona.dni or "-",
            Paragraph(persona.apellidos_nombres or "-", style_table),  # Se ajusta dentro de la celda
            Paragraph(persona.correo_personal or "-", style_table),
            persona.celular or "-",
            persona.periodo_inicio.strftime("%d/%m/%Y") if persona.periodo_inicio else "-",
            persona.periodo_fin.strftime("%d/%m/%Y") if persona.periodo_fin else "-",
            Paragraph(sede, style_table),
            es_planilla
        ])

    # Definir anchos din√°micos para evitar desbordamiento
    col_widths = [60, 120, 140, 80, 70, 70, 100, 60]

    # Crear la tabla
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003366")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f2f2f2")),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#e6e6e6")]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)

    return response


@login_required
def agregar_banco(request):
    if request.method == 'POST':
        nombre_banco = request.POST.get('nombre')

        if nombre_banco:
            Banco.objects.create(nombre=nombre_banco)
            messages.success(request, 'Banco agregado exitosamente.')
        else:
            messages.error(request, 'El nombre del banco es obligatorio.')

        # Redirige a la misma p√°gina desde donde se envi√≥ la solicitud
        return redirect(request.META.get('HTTP_REFERER', 'ingreso'))  # Si no hay REFERER, usa 'ingreso'

    return redirect('ingreso')




def prestamos(request):
    if request.method == 'POST':
        try:
            fecha_prestamo = request.POST.get('fecha_prestamo')
            fecha_vencimiento = request.POST.get('fecha_vencimiento')
            numero_cuotas = request.POST.get('numero_cuotas')
            proveedor_id = request.POST.get('proveedor')
            tea = request.POST.get('tea')
            banco_id = request.POST.get('banco')
            analista = request.POST.get('analista')
            monto = request.POST.get('monto')
            estado = request.POST.get('estado')
            local_id = request.POST.get('local')
            dia_pago = request.POST.get('dia_pago')  # Recibimos el d√≠a de pago
            cuota_actual = request.POST.get('cuota_actual') if estado == "proceso" else None
            monto_cuota = request.POST.get('monto_cuota')
            notas=request.POST.get('notas')
            # Validaci√≥n de datos
            if not all([fecha_prestamo, numero_cuotas, monto, estado, local_id]):
                messages.error(request, "Todos los campos son requeridos.")
                return redirect('crear_prestamo')

            numero_cuotas = int(numero_cuotas)
            tea = float(tea) if tea not in [None, ""] else 0.0  # Se asigna 0.0 si tea es nulo o vac√≠o
            monto = float(monto)
            cuota_actual = int(cuota_actual) if cuota_actual and cuota_actual.isdigit() else 1
            dia_pago = int(dia_pago)  # Convertir a entero

            # Obtener proveedor, banco y local si existen
            proveedor = Proveedor.objects.get(id=proveedor_id) if proveedor_id else None
            banco = Banco.objects.get(id=banco_id) if banco_id else None
            local = Local.objects.get(id=local_id) if local_id else None

            # Crear el pr√©stamo
            prestamo = Prestamo(
                fecha_prestamo=fecha_prestamo,
                fecha_vencimiento=fecha_vencimiento if fecha_vencimiento else None,  # Solo se guarda si tiene valor
                numero_cuotas=numero_cuotas,
                proveedor=proveedor,
                tea=tea,
                banco=banco,
                analista=analista,
                monto=monto,
                estado="proceso" if estado == "nuevo" else estado,  # Se asigna "proceso" si era "nuevo"
                cuota_actual=cuota_actual,
                local=local,
                dia_pago=dia_pago,  # Guardamos el d√≠a de pago
                monto_cuota=monto_cuota,
                notas=notas
            )
            prestamo.save()

            # Solo crear el ingreso si el estado es "nuevo"
            if estado == "nuevo":
                fondo = Fondo.objects.filter(nombre_fondo="Fondo Prestamo").first()

                ingreso = Ingreso.objects.create(
                    prestamo=prestamo,
                    fecha_registro=now().date(),
                    fecha_ingreso=now().date(),
                    importe=monto,
                    id_fondo=fondo,
                    metodo_pago="efectivo",
                    moneda="Soles",
                    local=local,  # Se asigna el local
                    observacion="Prestamo"
                )

            messages.success(request, "Pr√©stamo e ingreso creados exitosamente.")
            return redirect('prestamos')  # Redirigir a la lista de pr√©stamos

        except ValueError:
            messages.error(request, "Error en los datos ingresados. Verifica los valores num√©ricos.")
        except Proveedor.DoesNotExist:
            messages.error(request, "El proveedor seleccionado no existe.")
        except Banco.DoesNotExist:
            messages.error(request, "El banco seleccionado no existe.")
        except Local.DoesNotExist:
            messages.error(request, "El local seleccionado no existe.")
        except Exception as e:
            messages.error(request, f"Ocurri√≥ un error: {str(e)}")

        return redirect('prestamos')

    # Obtener proveedores, bancos y locales
    proveedores = Proveedor.objects.all()
    bancos = Banco.objects.all()
    locales = Local.objects.all()
    return render(request, 'crear_prestamos.html', {'proveedores': proveedores, 'bancos': bancos, 'locales': locales})

@login_required
def ingreso(request):
    if request.method == 'POST':
        fecha = request.POST.get('fecha')
        importe = request.POST.get('importe')
        id_fondo = request.POST.get('id_fondo')
        metodo_pago = request.POST.get('metodo_pago')
        moneda = request.POST.get('moneda')
        local_id = request.POST.get('local')
        observacion = request.POST.get('observacion')
        codigo_operacion = request.POST.get('codigo_operacion') if metodo_pago != 'efectivo' else None
        fecha_operacion  = request.POST.get('fecha_operacion') if metodo_pago != 'efectivo' else None
        banco_id = request.POST.get('banco_operacion')  # Obtener el banco seleccionado

        banco = Banco.objects.get(id=banco_id) if banco_id else None

        # Validaci√≥n b√°sica
        if not (fecha and importe and id_fondo and metodo_pago and moneda):
            messages.error(request, "Todos los campos obligatorios deben completarse.")
            return redirect('ingreso')

        # Guardar el ingreso
        fondo = Fondo.objects.get(id=id_fondo)
        local = Local.objects.get(id=local_id) if local_id else None

        ingreso = Ingreso(
            fecha_ingreso=fecha,
            importe=importe,
            id_fondo=fondo,
            metodo_pago=metodo_pago,
            moneda=moneda,
            local=local,
            observacion=observacion,
            codigo_operacion=codigo_operacion,  # Ya est√° seguro para el caso de transferencia
            fecha_operacion=fecha_operacion,
            banco=banco  # Guardar el banco seleccionado

        )
        ingreso.usuario_creador=request.user
        ingreso.save()
        messages.success(request, "Ingreso registrado correctamente.")
        return redirect('ingreso')

    # Contexto para renderizar el formulario
    fondos = Fondo.objects.all()
    locales = Local.objects.all()
    bancos = Banco.objects.all()
    return render(request, 'ingreso.html', {'fondos': fondos, 'locales': locales,'bancos': bancos})

from datetime import datetime

def comprobar_conceptos(tipo_comprobante, concepto_nivel_1, concepto_nivel_2, concepto_nivel_3):
    # Validaci√≥n jer√°rquica de los conceptos (excepto para tipo_comprobante "Requerimiento")
    if tipo_comprobante not in ['Requerimiento', 'Sin Requerimiento']:
        # Verifica si se proporcion√≥ concepto_nivel_1
        if not concepto_nivel_1:
            return JsonResponse({'error': 'Debe seleccionar un concepto de nivel 1.'}, status=400)

        # Intenta obtener el concepto de nivel 1
        try:
            concepto_1 = Concepto.objects.get(id=concepto_nivel_1)
        except Concepto.DoesNotExist:
            return JsonResponse({'error': 'El concepto de nivel 1 especificado no existe.'}, status=404)

        # Si el nivel 1 tiene conceptos hijos, validar el nivel 2
        conceptos_nivel_2 = Concepto.objects.filter(id_concepto_padre=concepto_1)
        if conceptos_nivel_2.exists():
            # Verifica si se proporcion√≥ concepto_nivel_2
            if not concepto_nivel_2:
                return JsonResponse({'error': 'Debe seleccionar un concepto de nivel 2 asociado al nivel 1.'}, status=400)

            try:
                concepto_2 = Concepto.objects.get(id=concepto_nivel_2)
            except Concepto.DoesNotExist:
                return JsonResponse({'error': 'El concepto de nivel 2 especificado no existe.'}, status=404)

            # Si el nivel 2 tiene conceptos hijos, validar el nivel 3
            conceptos_nivel_3 = Concepto.objects.filter(id_concepto_padre=concepto_2)
            if conceptos_nivel_3.exists():
                # Verifica si se proporcion√≥ concepto_nivel_3
                if not concepto_nivel_3:
                    return JsonResponse({'error': 'Debe seleccionar un concepto de nivel 3 asociado al nivel 2.'}, status=400)

                try:
                    concepto_3 = Concepto.objects.get(id=concepto_nivel_3)
                except Concepto.DoesNotExist:
                    return JsonResponse({'error': 'El concepto de nivel 3 especificado no existe.'}, status=404)
            else:
                concepto_3 = None  # No hay nivel 3 asociado
        else:
            concepto_2 = None  # No hay nivel 2 asociado
            concepto_3 = None  # No hay nivel 3 asociado
    else:
        # Si el tipo de comprobante es "Requerimiento", los conceptos pueden ser vac√≠os
        concepto_1 = None
        concepto_2 = None
        concepto_3 = None

    # Retornar los conceptos validados
    return {
        'concepto_nivel_1': concepto_1,
        'concepto_nivel_2': concepto_2,
        'concepto_nivel_3': concepto_3,
    }

@login_required
def gasto(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Captura de datos del request
            fecha = data.get('fecha')
            importe = data.get('importe')
            metodo_pago = data.get('metodo_pago')
            moneda = data.get('moneda')
            local_id = data.get('local')
            tipo_comprobante = data.get('tipo_comprobante')
            nombre_proveedor = data.get('nombre_proveedor')
            observacion = data.get('observacion')
            codigo_operacion = data.get('codigo_operacion') if metodo_pago != 'efectivo' else None
            fecha_operacion = data.get('fecha_operacion') if metodo_pago != 'efectivo' else None
            concepto_nivel_1 = data.get('concepto_nivel_1')
            concepto_nivel_2 = data.get('concepto_nivel_2')
            concepto_nivel_3 = data.get('concepto_nivel_3')
            numero_comprobante = data.get('num_comprobante')
            fecha_emision_comprobante = data.get('fecha_emision_comprobante') if tipo_comprobante  in ['RHE', 'Factura', 'Boleta','Nota','Proforma'] else None # Nuevo campo
            campo_area=data.get('campo_area')
            campo_mes=data.get('campo_mes') if tipo_comprobante  in ['Boleta de pago'] else None
            id_requerimiento = data.get('id_requerimiento')  # Nuevo campo
            num_requerimiento = data.get('num_requerimiento')  # Nuevo campo
            banco_id =data.get('banco_operacion') if metodo_pago != 'efectivo' else None
            # Buscar la instancia del banco si 'banco_id' no es None
            if banco_id:
                try:
                    banco = Banco.objects.get(id=banco_id)  # Buscar la instancia del banco por ID
                except Banco.DoesNotExist:
                    return JsonResponse({'error': 'El banco especificado no existe.'}, status=404)
            else:
                banco = None

            if fecha:
                try:
                    fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
                except ValueError:
                    return JsonResponse({'error': 'El valor de "fecha" tiene un formato inv√°lido. Deber√≠a estar en el formato YYYY-MM-DD.'}, status=400)

            # Validaci√≥n y conversi√≥n de fechas

            # Validaci√≥n b√°sica de campos obligatorios
            if not (fecha and importe and metodo_pago and moneda):
                return JsonResponse({'error': 'Todos los campos obligatorios deben completarse.'}, status=400)

            # Validaci√≥n adicional para tipo_comprobante espec√≠fico
            if tipo_comprobante in ['RHE', 'Factura', 'Boleta']:
                if not numero_comprobante or not fecha_emision_comprobante:
                    return JsonResponse({
                        'error': 'Los campos "N√∫mero de Comprobante" y "Fecha de Emisi√≥n del Comprobante" son obligatorios para el tipo de comprobante seleccionado.'
                    }, status=400)
            if tipo_comprobante == 'Boleta de pago':
                if not campo_mes:
                    return JsonResponse({
                        'error': 'El campo "Mes" es obligatorio cuando el tipo de comprobante es "Boleta de pago".'
                    }, status=400)
            # Validaci√≥n adicional para tipo_comprobante "Requerimiento"
            if tipo_comprobante == 'Requerimiento':
                if not id_requerimiento:
                    return JsonResponse({
                        'error': 'El campo "ID Requerimiento" es obligatorio cuando el tipo de comprobante es "Requerimiento".'
                    }, status=400)
            # Buscar el proveedor por nombre (asumido que es √∫nico)
            proveedor = Proveedor.objects.filter(id=nombre_proveedor).first()
            if not proveedor:
                return JsonResponse({'error': 'El proveedor especificado no existe.'}, status=404)

            # Obtener el local
            local = Local.objects.get(id=local_id) if local_id else None

            # Validar conceptos usando la funci√≥n
            resultado_conceptos = comprobar_conceptos(tipo_comprobante, concepto_nivel_1, concepto_nivel_2, concepto_nivel_3)

            if 'error' in resultado_conceptos:
                return JsonResponse({'error': 'Revise los niveles de los conceptos, falta llenar algunos campos.'}, status=400)

            # Extraer los conceptos validados
            concepto_1 = resultado_conceptos['concepto_nivel_1']
            concepto_2 = resultado_conceptos['concepto_nivel_2']
            concepto_3 = resultado_conceptos['concepto_nivel_3']

            # Crear y guardar la instancia del gasto
            gasto = Gasto(
                fecha_gasto=fecha,
                concepto_nivel_1=concepto_1,
                concepto_nivel_2=concepto_2,
                concepto_nivel_3=concepto_3,
                importe=importe,
                nombre_proveedor=proveedor,
                local=local,
                tipo_comprobante=tipo_comprobante,
                tipo_pago=metodo_pago,
                codigo_operacion=codigo_operacion,
                observacion=observacion,
                fecha_operacion=fecha_operacion,
                moneda=moneda,
                numero_comprobante=numero_comprobante,
                fecha_emision_comprobante=fecha_emision_comprobante,
                campo_area=campo_area,
                campo_mes=campo_mes,
                id_requerimiento=id_requerimiento if tipo_comprobante == 'Requerimiento' else None ,
                num_requerimiento=num_requerimiento if tipo_comprobante == 'Requerimiento' else None,
                banco=banco  # Asignamos la instancia del banco si existe


            )
            gasto.usuario_creador = request.user

            gasto.save()

            return JsonResponse({'success': 'Gasto registrado correctamente.'}, status=200)

        except Local.DoesNotExist:
            return JsonResponse({'error': 'El local seleccionado no existe.'}, status=404)
        except Concepto.DoesNotExist:
            return JsonResponse({'error': 'Alguno de los conceptos seleccionados no existe.'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Ocurri√≥ un error inesperado: {str(e)}'}, status=500)

    nivel_1_conceptos = Concepto.objects.filter(nivel=1)
    nivel_2_conceptos = Concepto.objects.filter(nivel=2)
    nivel_3_conceptos = Concepto.objects.filter(nivel=3)
    fondos = Fondo.objects.all()
    locales = Local.objects.all()
    bancos = Banco.objects.all()
    proveedores = Proveedor.objects.all()  # Obtener todos los proveedores
    nivel_1_conceptos_json = json.dumps(list(nivel_1_conceptos.values_list('id', 'concepto_nombre')), default=str)

    return render(request, 'gasto.html', {
        'nivel_1_conceptos_json': nivel_1_conceptos_json,
        'fondos': fondos,
        'locales': locales,
        'proveedores': proveedores,  # Pasar los proveedores al contexto
        'nivel_1_conceptos': nivel_1_conceptos,
        'nivel_2_conceptos': nivel_2_conceptos,
        'nivel_3_conceptos': nivel_3_conceptos,
        'bancos':bancos
    })
@csrf_exempt
def finalizar_cuota(request, prestamo_id):
    if request.method == "POST":
        try:
            prestamo = Prestamo.objects.get(id=prestamo_id)
            if prestamo.cuota_actual < prestamo.numero_cuotas:
                prestamo.cuota_actual += 1
                prestamo.save()
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'error': 'El pr√©stamo ya est√° finalizado.'}, status=400)
        except Prestamo.DoesNotExist:
            return JsonResponse({'error': 'Pr√©stamo no encontrado.'}, status=404)

    return JsonResponse({'error': 'M√©todo no permitido.'}, status=405)
def ver_prestamos(request):
    # Obtener todos los pr√©stamos
    prestamos = Prestamo.objects.all()
    bancos = Banco.objects.all()
    return render(request, 'ver_prestamos.html', {'prestamos_data': prestamos ,'bancos': bancos  })

def ver_pagos(request, prestamo_id):
    try:
        prestamo = Prestamo.objects.get(id=prestamo_id)
        pagos = Pago.objects.filter(prestamo=prestamo, cuota=prestamo.cuota_actual).values(
            'fecha_pago', 'monto_pagado', 'cuota', 'notas'
        )
        return JsonResponse({'pagos': list(pagos)})
    except Prestamo.DoesNotExist:
        return JsonResponse({'error': 'Pr√©stamo no encontrado'}, status=404)

def ficha_ingreso_view(request):
    bancos = Banco.objects.all()
    locales = Local.objects.all()

    return render(request, "ficha_ingreso.html",{'bancos':bancos,'locales':locales})

def realizar_pago(request):
    if request.method == 'POST':
        prestamo_id = request.POST.get('prestamo_id', '').strip()
        monto_pagado = request.POST.get('monto_pagado', '').strip()
        fecha_pago = request.POST.get('fecha_pago', '').strip()
        # Campos para medio de pago
        medio_pago = request.POST.get('medio_pago', '').strip()  # "efectivo", "deposito" o "transferencia"
        banco_id = request.POST.get('banco', '').strip()
        codigo_operacion = request.POST.get('codigo_operacion', '').strip()
        fecha_operacion = request.POST.get('fecha_operacion', '').strip()
        nota = request.POST.get('nota', '').strip()  # Captura la nota ingresada

        if not prestamo_id.isdigit():
            messages.error(request, "‚ùå ID de pr√©stamo inv√°lido.")
            return redirect('ver_prestamos')

        if not monto_pagado:
            messages.error(request, "‚ùå Debe ingresar un monto de pago.")
            return redirect('ver_prestamos')

        try:
            monto_pagado = Decimal(monto_pagado)
            if monto_pagado <= 0:
                raise ValueError
        except:
            messages.error(request, "‚ùå Monto de pago no v√°lido.")
            return redirect('ver_prestamos')

        prestamo = get_object_or_404(Prestamo, id=int(prestamo_id))
        cuota_actual = prestamo.cuota_actual
        monto_cuota = prestamo.monto_cuota
        numero_cuotas = prestamo.numero_cuotas

        if cuota_actual > numero_cuotas:
            messages.warning(request, "‚ö†Ô∏è Este pr√©stamo ya ha sido cancelado.")
            return redirect('ver_prestamos')


        Pago.objects.create(
            prestamo=prestamo,
            cuota=cuota_actual,
            monto_pagado=monto_pagado,
            fecha_pago=fecha_pago,
            notas=nota
        )
        prestamo.save()
            # Configurar tipo de pago seg√∫n el medio seleccionado
        if medio_pago == "efectivo":
            tipo_pago = "efectivo"
            tipo_comprobante = "Sin comprobante"
            # Para efectivo, los campos de banco, c√≥digo y fecha de operaci√≥n se ignoran
            banco_id = None
            codigo_operacion = ""
            fecha_operacion = ""
        elif medio_pago in ["deposito", "transferencia"]:
            tipo_pago = "transferencia"
            # Se puede diferenciar si es dep√≥sito o transferencia, si se requiere:
            tipo_comprobante = "Deposito en cuenta" if medio_pago == "Deposito en cuenta" else "Transferencia"
        else:
            # Valor por defecto en caso de no seleccionar un medio adecuado
            tipo_pago = "efectivo"
            tipo_comprobante = "Sin comprobante"
            banco_id = None
            codigo_operacion = ""
            fecha_operacion = ""

        # Obtener o crear el concepto de nivel 1 (PAGO DE PR√âSTAMOS)
        concepto_nivel_1, created = Concepto.objects.get_or_create(
                concepto_nombre="PAGO DE PRESTAMOS",
                nivel=1,
                id_concepto_padre=None  # Nivel 1 no tiene padre
            )

            # Obtener o crear el concepto de nivel 2 (PR√âSTAMOS DE TERCEROS) con nivel 1 como padre
        concepto_nivel_2, created = Concepto.objects.get_or_create(
                concepto_nombre="PRESTAMOS DE TERCEROS",
                nivel=2,
                id_concepto_padre=concepto_nivel_1
            )
        # Crear el gasto asociado
        Gasto.objects.create(
            usuario_creador=request.user,
            fecha_gasto=now().date(),
            prestamo=prestamo,
            banco_id=banco_id,  # Solo tendr√° valor si el medio de pago es dep√≥sito o transferencia
            fecha_operacion = fecha_operacion if medio_pago != "efectivo" else None,
            codigo_operacion=codigo_operacion,
            importe=monto_pagado,
            moneda="Soles",
            tipo_pago=tipo_pago,
            tipo_comprobante=tipo_comprobante,
            nombre_proveedor=prestamo.proveedor,  # Asociar el proveedor del pr√©stamo
            local=prestamo.local,  # Asociar el local del pr√©stamo
            observacion=nota,  # Se guarda la nota en el gasto tambi√©n
            concepto_nivel_1=concepto_nivel_1,
            concepto_nivel_2=concepto_nivel_2,
            concepto_nivel_3=None
        )
        messages.success(request, "‚úÖ Pago registrado ,se creo un Gasto !  üéâ")
        return redirect('ver_prestamos')

    return HttpResponse(status=405)

def get_nivel_2_conceptos(request):
    id_concepto_padre = request.GET.get('id_concepto_padre')
    conceptos_nivel_2 = Concepto.objects.filter(id_concepto_padre=id_concepto_padre)
    return JsonResponse(list(conceptos_nivel_2.values('id', 'concepto_nombre')), safe=False)

def get_nivel_3_conceptos(request):
    id_concepto_padre = request.GET.get('id_concepto_padre')
    conceptos_nivel_3 = Concepto.objects.filter(id_concepto_padre=id_concepto_padre)
    return JsonResponse(list(conceptos_nivel_3.values('id', 'concepto_nombre')), safe=False)



from django.shortcuts import get_object_or_404
@login_required
def rendicion(request):
    # Obtener la fecha actual
    today = date.today().isoformat()

    if request.user.is_superuser or request.user.is_staff:
        # Si el usuario es superuser o staff, se env√≠an todos los gastos que cumplen las condiciones
        gastos_requerimientos = Gasto.objects.filter(
            Q(tipo_comprobante="Requerimiento") | Q(tipo_comprobante="Sin Requerimiento"),
            rendido=False,
            gasto_origen__isnull=True  # Filtrar por gasto_origen nulo
        )
    else:
        # Filtrar por usuario autenticado
        gastos_requerimientos = Gasto.objects.filter(
            tipo_comprobante="Requerimiento",
            rendido=False,
            gasto_origen__isnull=True,  # Filtrar por gasto_origen nulo
            usuario_creador=request.user
        )

    # Renderizar la plantilla con los gastos filtrados
    return render(request, "rendicion.html", {"today": today, "gastos": gastos_requerimientos})











